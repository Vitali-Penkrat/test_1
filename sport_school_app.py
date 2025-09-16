# -*- coding: utf-8 -*-
"""
Sports DB — шаг 7+: поиск + скроллы + пагинация + подсветка мест 1–3.

Запуск: py sports_app_step7.py
"""

import os, io, csv, sqlite3, datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime as dt

# --- XLSX поддержка (опционально)
try:
    import openpyxl  # type: ignore
    HAS_XLSX = True
except Exception:
    HAS_XLSX = False

APP_TITLE = "Sports DB — шаг 7+ (Поиск/Пагинация/Карточки/Подсветка)"
APP_SIZE  = (1180, 780)

LEVELS = ["", "Район", "Область", "Республика", "Международные"]  # '' = все
LINES  = ["", "Образование", "Спорт"]
SPORTS = ["", "Ориентирование", "Туризм", "Спартакиада (разное)"]
MEDALS = ["", "gold", "silver", "bronze"]

DB_PATH = "sports.db"

SCHEMA_SQL = r"""
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS coaches (
    coach_id INTEGER PRIMARY KEY,
    fio TEXT NOT NULL,
    phone TEXT
);

CREATE TABLE IF NOT EXISTS groups (
    group_id INTEGER PRIMARY KEY,
    name TEXT NOT NULL,
    sport TEXT NOT NULL,
    coach_id INTEGER,
    FOREIGN KEY (coach_id) REFERENCES coaches(coach_id) ON DELETE RESTRICT
);
CREATE UNIQUE INDEX IF NOT EXISTS uq_groups_name_sport ON groups(name, sport);

CREATE TABLE IF NOT EXISTS persons (
    person_id INTEGER PRIMARY KEY,
    last_name   TEXT NOT NULL,
    first_name  TEXT NOT NULL,
    birthdate   TEXT,        -- YYYY-MM-DD
    address     TEXT,
    phone       TEXT,
    group_id    INTEGER,
    FOREIGN KEY (group_id) REFERENCES groups(group_id) ON DELETE SET NULL
);

CREATE TABLE IF NOT EXISTS events (
    event_id INTEGER PRIMARY KEY,
    name TEXT NOT NULL,
    date TEXT NOT NULL,   -- YYYY-MM-DD
    level TEXT NOT NULL,
    line  TEXT NOT NULL,
    sport TEXT NOT NULL,
    location TEXT,
    total_count INTEGER,
    UNIQUE(name, date, location)
);

CREATE TABLE IF NOT EXISTS results (
    result_id INTEGER PRIMARY KEY,
    event_id   INTEGER NOT NULL,
    person_id  INTEGER NOT NULL,
    category   TEXT NOT NULL DEFAULT '',
    place      INTEGER,
    medal      TEXT,
    note       TEXT,
    UNIQUE(event_id, person_id, category),
    FOREIGN KEY (event_id)  REFERENCES events(event_id)   ON DELETE CASCADE,
    FOREIGN KEY (person_id) REFERENCES persons(person_id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_results_event  ON results(event_id);
CREATE INDEX IF NOT EXISTS idx_results_person ON results(person_id);
"""

# ------------ Хранилище ------------
class Store:
    def __init__(self, db_path=DB_PATH):
        self.conn = sqlite3.connect(db_path)
        self.conn.execute("PRAGMA foreign_keys = ON;")
        self.conn.row_factory = sqlite3.Row
        self.conn.executescript(SCHEMA_SQL)
        self.conn.commit()

    # helpers
    def _fetchall(self, q, a=()):
        return [dict(r) for r in self.conn.execute(q, a).fetchall()]
    def _fetchone(self, q, a=()):
        r = self.conn.execute(q, a).fetchone()
        return dict(r) if r else None

    # --- coaches
    def add_coach(self, fio, phone):
        self.conn.execute("INSERT INTO coaches(fio,phone) VALUES(?,?)", (fio, phone or None)); self.conn.commit()
    def list_coaches(self):
        return self._fetchall("SELECT coach_id, fio, COALESCE(phone,'') AS phone FROM coaches ORDER BY coach_id")
    def edit_coach(self, cid, fio, phone):
        self.conn.execute("UPDATE coaches SET fio=?, phone=? WHERE coach_id=?", (fio, phone or None, cid)); self.conn.commit()
    def can_delete_coach(self, cid):
        return self._fetchone("SELECT 1 FROM groups WHERE coach_id=? LIMIT 1", (cid,)) is None
    def delete_coach(self, cid):
        self.conn.execute("DELETE FROM coaches WHERE coach_id=?", (cid,)); self.conn.commit()

    # --- groups
    def add_group(self, name, sport, coach_id):
        self.conn.execute("INSERT INTO groups(name,sport,coach_id) VALUES(?,?,?)", (name, sport, coach_id)); self.conn.commit()
    def list_groups(self):
        q = """
        SELECT g.group_id, g.name, g.sport, COALESCE(c.fio,'—') AS coach,
               (SELECT COUNT(*) FROM persons p WHERE p.group_id=g.group_id) AS members
        FROM groups g LEFT JOIN coaches c ON c.coach_id=g.coach_id ORDER BY g.group_id"""
        return self._fetchall(q)
    def edit_group(self, gid, name, sport, coach_id):
        self.conn.execute("UPDATE groups SET name=?, sport=?, coach_id=? WHERE group_id=?", (name, sport, coach_id, gid)); self.conn.commit()
    def can_delete_group(self, gid):
        return self._fetchone("SELECT 1 FROM persons WHERE group_id=? LIMIT 1", (gid,)) is None
    def delete_group(self, gid):
        self.conn.execute("DELETE FROM groups WHERE group_id=?", (gid,)); self.conn.commit()
    def group_info(self, gid):
        g = self._fetchone("SELECT * FROM groups WHERE group_id=?", (gid,))
        members = self._fetchall("""SELECT person_id, last_name||' '||first_name AS fio,
                                    COALESCE(birthdate,'') AS birthdate, COALESCE(phone,'') AS phone
                                    FROM persons WHERE group_id=? ORDER BY last_name, first_name""",(gid,))
        return g, members

    # --- persons
    def add_person(self, last, first, birthdate, address, phone, group_id):
        self.conn.execute("""INSERT INTO persons(last_name,first_name,birthdate,address,phone,group_id)
                             VALUES(?,?,?,?,?,?)""",(last, first, birthdate or None, address or None, phone or None, group_id))
        self.conn.commit()
    def list_persons(self):
        q = """
        SELECT p.person_id, p.last_name||' '||p.first_name AS fio, COALESCE(p.birthdate,'') AS birthdate,
               COALESCE(g.name,'—') AS gname, COALESCE(c.fio,'—') AS coach,
               COALESCE(p.phone,'') AS phone, COALESCE(p.address,'') AS address
        FROM persons p
        LEFT JOIN groups g  ON g.group_id=p.group_id
        LEFT JOIN coaches c ON c.coach_id=g.coach_id
        ORDER BY p.person_id"""
        return self._fetchall(q)
    def person_raw(self, pid):
        return self._fetchone("SELECT * FROM persons WHERE person_id=?", (pid,))
    def edit_person(self, pid, last, first, birthdate, address, phone, group_id):
        self.conn.execute("""UPDATE persons SET last_name=?, first_name=?, birthdate=?, address=?, phone=?, group_id=?
                             WHERE person_id=?""", (last, first, birthdate or None, address or None, phone or None, group_id, pid))
        self.conn.commit()
    def can_delete_person(self, pid):
        return self._fetchone("SELECT 1 FROM results WHERE person_id=? LIMIT 1", (pid,)) is None
    def delete_person(self, pid):
        self.conn.execute("DELETE FROM persons WHERE person_id=?", (pid,)); self.conn.commit()

    # --- events
    def add_event(self, name, date, level, line, sport, location, total):
        self.conn.execute("""INSERT INTO events(name,date,level,line,sport,location,total_count)
                             VALUES(?,?,?,?,?,?,?)""",(name, date, level, line, sport, location or None, total))
        self.conn.commit()
    def list_events(self):
        q = """
        SELECT e.event_id, e.date, e.name, e.level, e.line, e.sport,
               COALESCE(e.location,'') AS location, COALESCE(e.total_count,'') AS total_count,
               (SELECT COUNT(DISTINCT person_id) FROM results r WHERE r.event_id=e.event_id) AS ours
        FROM events e ORDER BY e.date DESC, e.event_id DESC"""
        return self._fetchall(q)
    def event_raw(self, eid):
        return self._fetchone("SELECT * FROM events WHERE event_id=?", (eid,))
    def edit_event(self, eid, name, date, level, line, sport, location, total):
        self.conn.execute("""UPDATE events SET name=?, date=?, level=?, line=?, sport=?, location=?, total_count=?
                             WHERE event_id=?""",(name, date, level, line, sport, location or None, total, eid))
        self.conn.commit()
    def can_delete_event(self, eid):
        return self._fetchone("SELECT 1 FROM results WHERE event_id=? LIMIT 1", (eid,)) is None
    def delete_event(self, eid):
        self.conn.execute("DELETE FROM events WHERE event_id=?", (eid,)); self.conn.commit()

    # --- results
    def add_result(self, event_id, person_id, category, place, medal, note):
        category = category or ''
        self.conn.execute("""INSERT OR REPLACE INTO results(event_id, person_id, category, place, medal, note)
                             VALUES(?,?,?,?,?,?)""",(event_id, person_id, category, place, medal or "", note or None))
        self.conn.commit()
    def list_results(self):
        q = """
        SELECT r.result_id, e.date, e.name AS event_name, p.last_name||' '||p.first_name AS fio,
               r.category, COALESCE(r.place,'') AS place, COALESCE(r.medal,'') AS medal, COALESCE(r.note,'') AS note
        FROM results r JOIN events e ON e.event_id=r.event_id JOIN persons p ON p.person_id=r.person_id
        ORDER BY e.date DESC, r.result_id DESC"""
        return self._fetchall(q)
    def result_raw(self, rid):
        return self._fetchone("SELECT * FROM results WHERE result_id=?", (rid,))
    def edit_result(self, rid, event_id, person_id, category, place, medal, note):
        category = category or ''
        self.conn.execute("""UPDATE results SET event_id=?, person_id=?, category=?, place=?, medal=?, note=?
                             WHERE result_id=?""",(event_id, person_id, category, place, medal or "", note or None, rid))
        self.conn.commit()
    def delete_result(self, rid):
        self.conn.execute("DELETE FROM results WHERE result_id=?", (rid,)); self.conn.commit()

    # --- общие помощники фильтра для отчётов
    def _where_events(self, flt):
        cond, par = [], []
        if flt.get("date_from"):
            cond.append("date >= ?"); par.append(flt["date_from"])
        if flt.get("date_to"):
            cond.append("date <= ?"); par.append(flt["date_to"])
        if flt.get("sport"):
            cond.append("sport = ?"); par.append(flt["sport"])
        if flt.get("line"):
            cond.append("line = ?"); par.append(flt["line"])
        if flt.get("level"):
            cond.append("level = ?"); par.append(flt["level"])
        where = "WHERE " + " AND ".join(cond) if cond else ""
        return where, par

    def _where_results_join(self, flt):
        cond, par = [], []
        if flt.get("date_from"):
            cond.append("e.date >= ?"); par.append(flt["date_from"])
        if flt.get("date_to"):
            cond.append("e.date <= ?"); par.append(flt["date_to"])
        if flt.get("sport"):
            cond.append("e.sport = ?"); par.append(flt["sport"])
        if flt.get("line"):
            cond.append("e.line = ?"); par.append(flt["line"])
        if flt.get("level"):
            cond.append("e.level = ?"); par.append(flt["level"])
        where = "WHERE " + " AND ".join(cond) if cond else ""
        return where, par

    # --- отчёты
    def medals_summary(self, flt):
        where, par = self._where_results_join(flt)
        q = f"""SELECT 
              SUM(CASE WHEN r.medal='gold'   THEN 1 ELSE 0 END) AS g,
              SUM(CASE WHEN r.medal='silver' THEN 1 ELSE 0 END) AS s,
              SUM(CASE WHEN r.medal='bronze' THEN 1 ELSE 0 END) AS b
            FROM results r JOIN events e ON e.event_id=r.event_id {where}"""
        row = self._fetchone(q, par) or {"g":0,"s":0,"b":0}
        return int(row["g"] or 0), int(row["s"] or 0), int(row["b"] or 0)

    def events_breakdown(self, flt):
        where, par = self._where_events(flt)
        by_level = {k:0 for k in LEVELS if k}
        by_line  = {k:0 for k in LINES if k}
        for r in self._fetchall(f"SELECT level, line FROM events {where}", par):
            if r["level"] in by_level: by_level[r["level"]] += 1
            if r["line"]  in by_line:  by_line[r["line"]]  += 1
        return by_level, by_line

    def medals_by_coach(self, flt):
        where, par = self._where_results_join(flt)
        q = f"""
        SELECT c.coach_id, c.fio,
               SUM(CASE WHEN r.medal='gold'   THEN 1 ELSE 0 END) AS g,
               SUM(CASE WHEN r.medal='silver' THEN 1 ELSE 0 END) AS s,
               SUM(CASE WHEN r.medal='bronze' THEN 1 ELSE 0 END) AS b,
               COUNT(*)                           AS starts,
               COUNT(DISTINCT e.event_id)         AS events,
               COUNT(DISTINCT p.person_id)        AS athletes
        FROM results r
        JOIN events  e ON e.event_id=r.event_id
        JOIN persons p ON p.person_id=r.person_id
        JOIN groups  g ON g.group_id=p.group_id
        JOIN coaches c ON c.coach_id=g.coach_id
        {where}
        GROUP BY c.coach_id, c.fio"""
        return self._fetchall(q, par)

    # ---- для карточек ----
    def person_report(self, pid, flt):
        where, par = self._where_results_join(flt)
        par = [pid] + par
        q = f"""
        SELECT e.date, e.name, e.level, e.line, e.sport, r.category, r.place, r.medal, COALESCE(r.note,'') AS note
        FROM results r JOIN events e ON e.event_id=r.event_id
        WHERE r.person_id=? {"AND "+ where[6:] if where else ""}
        ORDER BY e.date DESC"""
        rows = self._fetchall(q, par)
        return rows

    def person_summary(self, pid, flt):
        rows = self.person_report(pid, flt)
        g = sum(1 for r in rows if r["medal"] == "gold")
        s = sum(1 for r in rows if r["medal"] == "silver")
        b = sum(1 for r in rows if r["medal"] == "bronze")
        prize = sum(1 for r in rows if r["medal"] in ("gold","silver","bronze") or (r["place"] and int(r["place"])<=3))
        return {"starts": len(rows), "gold": g, "silver": s, "bronze": b, "prize": prize}

    def coach_results(self, coach_id, flt):
        where, par = self._where_results_join(flt)
        par = [coach_id] + par
        q = f"""
        SELECT e.date, e.name AS event_name, 
               p.last_name||' '||p.first_name AS fio,
               r.category, r.place, r.medal, COALESCE(r.note,'') AS note
        FROM results r
        JOIN events  e ON e.event_id=r.event_id
        JOIN persons p ON p.person_id=r.person_id
        JOIN groups  g ON g.group_id=p.group_id
        WHERE g.coach_id=? {"AND "+ where[6:] if where else ""}
        ORDER BY e.date DESC"""
        return self._fetchall(q, par)

    def coach_summary(self, coach_id, flt):
        where, par = self._where_results_join(flt)
        par = [coach_id] + par
        q = f"""
        SELECT 
          SUM(CASE WHEN r.medal='gold'   THEN 1 ELSE 0 END) AS g,
          SUM(CASE WHEN r.medal='silver' THEN 1 ELSE 0 END) AS s,
          SUM(CASE WHEN r.medal='bronze' THEN 1 ELSE 0 END) AS b,
          COUNT(*)                           AS starts,
          COUNT(DISTINCT e.event_id)         AS events,
          COUNT(DISTINCT p.person_id)        AS athletes
        FROM results r
        JOIN events  e ON e.event_id=r.event_id
        JOIN persons p ON p.person_id=r.person_id
        JOIN groups  g ON g.group_id=p.group_id
        WHERE g.coach_id=? {"AND "+ where[6:] if where else ""}"""
        row = self._fetchone(q, par) or {}
        for k in ("g","s","b","starts","events","athletes"):
            row[k] = int(row.get(k) or 0)
        return row

    def group_report(self, gid, flt):
        where, par = self._where_results_join(flt)
        par = [gid] + par
        q = f"""
        SELECT e.date, e.name, pers.last_name||' '||pers.first_name AS fio, r.category, r.place, r.medal
        FROM results r
        JOIN events  e   ON e.event_id=r.event_id
        JOIN persons pers ON pers.person_id=r.person_id
        WHERE pers.group_id=? {"AND "+ where[6:] if where else ""}
        ORDER BY e.date DESC"""
        rows = self._fetchall(q, par)
        return rows

    def yearly_dynamics(self, flt):
        where_e, par_e = self._where_events(flt)
        q1 = f"""SELECT substr(e.date,1,4) AS y, COUNT(DISTINCT e.event_id) AS events
                 FROM events e
                 WHERE EXISTS (SELECT 1 FROM results r WHERE r.event_id=e.event_id)
                 {"AND "+ where_e[6:] if where_e else ""}
                 GROUP BY y ORDER BY y"""
        starts = {r["y"]: int(r["events"]) for r in self._fetchall(q1, par_e)}

        where_r, par_r = self._where_results_join(flt)
        q2 = f"""SELECT substr(e.date,1,4) AS y,
                        SUM(CASE WHEN r.medal='gold'   THEN 1 ELSE 0 END) AS g,
                        SUM(CASE WHEN r.medal='silver' THEN 1 ELSE 0 END) AS s,
                        SUM(CASE WHEN r.medal='bronze' THEN 1 ELSE 0 END) AS b
                 FROM results r JOIN events e ON e.event_id=r.event_id
                 {where_r}
                 GROUP BY y ORDER BY y"""
        medals = {r["y"]: (int(r["g"] or 0), int(r["s"] or 0), int(r["b"] or 0)) for r in self._fetchall(q2, par_r)}

        years = sorted(set(starts.keys()) | set(medals.keys()))
        rows = []
        for y in years:
            g,s,b = medals.get(y,(0,0,0))
            rows.append({"year": y, "events": starts.get(y,0), "gold": g, "silver": s, "bronze": b, "total_medals": g+s+b})
        return rows

    def event_results(self, event_id):
        q = """
        SELECT e.date, e.name AS event_name, e.level, e.line, e.sport,
               p.last_name||' '||p.first_name AS fio,
               COALESCE(gr.name,'—') AS gname,
               COALESCE(c.fio,'—')  AS coach,
               r.category, r.place, r.medal, COALESCE(r.note,'') AS note
        FROM results r
        JOIN events  e  ON e.event_id=r.event_id
        JOIN persons p  ON p.person_id=r.person_id
        LEFT JOIN groups gr ON gr.group_id=p.group_id
        LEFT JOIN coaches c ON c.coach_id=gr.coach_id
        WHERE e.event_id=?
        ORDER BY COALESCE(r.place, 999999), r.medal DESC, fio"""
        return self._fetchall(q, (event_id,))

# ------------ UI ------------
class App(tk.Tk):
    # --------- вспомогательные мини-компоненты (скроллы/пагинация/поиск) ----------
    class PagedSearchTable:
        def __init__(self, app, parent, columns, widths, get_rows_fn, apply_tags_fn=None, search_label="Поиск"):
            self.app = app
            self.columns = columns
            self.widths  = widths
            self.get_rows_fn = get_rows_fn
            self.apply_tags_fn = apply_tags_fn

            # верхняя панель: поиск + пагинация
            top = ttk.Frame(parent); top.pack(fill="x", padx=8, pady=(0,6))
            ttk.Label(top, text=f"{search_label}:").pack(side="left")
            self.var_q = tk.StringVar()
            ent = ttk.Entry(top, textvariable=self.var_q, width=36)
            ent.pack(side="left", padx=(4,10))
            ent.bind("<KeyRelease>", self._on_search)

            ttk.Label(top, text="На странице:").pack(side="left")
            self.var_page_size = tk.StringVar(value="50")
            cb = ttk.Combobox(top, textvariable=self.var_page_size, width=5, values=["10","20","50","100","200"])
            cb.pack(side="left", padx=(4,10))
            cb.bind("<<ComboboxSelected>>", lambda e: self._goto_page(0))

            self.btn_prev = ttk.Button(top, text="⟨ Назад", width=10, command=self.prev_page)
            self.btn_next = ttk.Button(top, text="Вперёд ⟩", width=10, command=self.next_page)
            self.btn_prev.pack(side="right")
            self.btn_next.pack(side="right", padx=(0,6))
            self.lbl_info = ttk.Label(top, text="—")
            self.lbl_info.pack(side="right", padx=(0,12))

            # дерево со скроллами
            frame = ttk.Frame(parent); frame.pack(fill="both", expand=True, padx=8, pady=4)
            tree = ttk.Treeview(frame, show="headings", columns=self.columns)
            for c,w in zip(self.columns, self.widths):
                tree.heading(c,text=c); tree.column(c,width=w,anchor="w")
            ysb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            xsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
            tree.grid(row=0, column=0, sticky="nsew")
            ysb.grid(row=0, column=1, sticky="ns")
            xsb.grid(row=1, column=0, sticky="ew")
            frame.rowconfigure(0, weight=1)
            frame.columnconfigure(0, weight=1)
            self.tree = tree

            self._all_rows = []
            self._filtered = []
            self._page = 0
            self.refresh()

        def refresh(self):
            self._all_rows = self.get_rows_fn()
            self._apply_filter()
            self._goto_page(0)

        def _apply_filter(self):
            q = (self.var_q.get() or "").strip().lower()
            if not q:
                self._filtered = list(self._all_rows)
                return
            def row_match(row):
                for v in row:
                    if q in str(v).lower():
                        return True
                return False
            self._filtered = [r for r in self._all_rows if row_match(r)]

        def _goto_page(self, page_idx):
            try:
                size = int(self.var_page_size.get())
                if size <= 0: size = 50
            except:
                size = 50
            n = len(self._filtered)
            max_page = (max(n-1,0)) // size
            self._page = max(0, min(page_idx, max_page))
            start = self._page * size
            end   = start + size
            rows = self._filtered[start:end]

            for i in self.tree.get_children():
                self.tree.delete(i)
            for r in rows:
                iid = self.tree.insert("", "end", values=r)
                if self.apply_tags_fn:
                    self.apply_tags_fn(self.tree, iid, r)

            self.lbl_info.config(text=f"Стр. {self._page+1}/{max_page+1} • всего: {n}")
            self.btn_prev.config(state=("normal" if self._page>0 else "disabled"))
            self.btn_next.config(state=("normal" if self._page<max_page else "disabled"))

        def next_page(self):
            self._goto_page(self._page + 1)
        def prev_page(self):
            self._goto_page(self._page - 1)
        def _on_search(self, *_):
            self._apply_filter()
            self._goto_page(0)

    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self._center(*APP_SIZE)
        self.store = Store()
        self._make_style()

        self.nb = ttk.Notebook(self); self.nb.pack(fill="both", expand=True)

        self._tab_persons()
        self._tab_coaches()
        self._tab_groups()
        self._tab_events()
        self._tab_results()
        self._tab_reports()
        self._tab_io()   # импорт/экспорт

    # helpers
    def _center(self, w, h):
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        x, y = (sw - w)//2, (sh - h)//3
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.minsize(1000, 660)

    def _make_style(self):
        style = ttk.Style()
        for theme in ("clam","default","classic"):
            try:
                style.theme_use(theme)
                break
            except:
                pass

    # ---- подсветка: медали + места ----
    def _init_medal_tags(self, tree):
        tree.tag_configure("medal_gold",  foreground="#a07800")
        tree.tag_configure("medal_silver", foreground="#707b8b")
        tree.tag_configure("medal_bronze", foreground="#8a572a")
    def _init_place_tags(self, tree):
        tree.tag_configure("place_1", background="#fff4ce")  # мягкий жёлтый
        tree.tag_configure("place_2", background="#eeeeee")  # светло-серый
        tree.tag_configure("place_3", background="#f2e2d5")  # мягкий бронзовый
    def _init_all_tags(self, tree):
        self._init_medal_tags(tree); self._init_place_tags(tree)

    def _add_tag(self, tree, iid, tag):
        if not tag: return
        cur = tree.item(iid, "tags") or ()
        if tag not in cur:
            tree.item(iid, tags=cur + (tag,))

    def _apply_medal_tag(self, tree, iid, medal):
        if medal == "gold":
            self._add_tag(tree, iid, "medal_gold")
        elif medal == "silver":
            self._add_tag(tree, iid, "medal_silver")
        elif medal == "bronze":
            self._add_tag(tree, iid, "medal_bronze")

    def _apply_place_tag(self, tree, iid, place):
        try:
            p = int(place)
        except:
            return
        if p == 1:
            self._add_tag(tree, iid, "place_1")
        elif p == 2:
            self._add_tag(tree, iid, "place_2")
        elif p == 3:
            self._add_tag(tree, iid, "place_3")

    def _fill_tree(self, tree, rows):
        for i in tree.get_children(): tree.delete(i)
        for r in rows: tree.insert("", "end", values=r)

    def _id_label(self, id_, label): return f"{id_} | {label}"
    def _parse_id(self, value):
        try: return int(str(value).split("|",1)[0].strip())
        except: return None
    def _timestamp(self): return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    def _selected_id(self, tree):
        sel=tree.selection(); return int(tree.item(sel[0])["values"][0]) if sel else None

    # -------- Участники --------
    def _tab_persons(self):
        f = ttk.Frame(self.nb); self.nb.add(f, text="Участники")
        form = ttk.LabelFrame(f, text="Добавить участника"); form.pack(fill="x", padx=8, pady=8)
        self.p_last=tk.Entry(form, width=20); self.p_first=tk.Entry(form, width=20); self.p_birth=tk.Entry(form, width=12)
        self.p_addr=tk.Entry(form, width=40); self.p_phone=tk.Entry(form, width=16)
        self.p_group=ttk.Combobox(form, width=40, values=self._group_options()); self.p_group.set("")
        ttk.Label(form, text="Фамилия").grid(row=0,column=0,sticky="w"); self.p_last.grid(row=0,column=1,padx=4)
        ttk.Label(form, text="Имя").grid(row=0,column=2,sticky="w"); self.p_first.grid(row=0,column=3,padx=4)
        ttk.Label(form, text="Дата рождения (YYYY-MM-DD)").grid(row=0,column=4,sticky="w"); self.p_birth.grid(row=0,column=5,padx=4)
        ttk.Label(form, text="Адрес").grid(row=1,column=0,sticky="w"); self.p_addr.grid(row=1,column=1,columnspan=3,sticky="we",padx=4)
        ttk.Label(form, text="Телефон").grid(row=1,column=4,sticky="w"); self.p_phone.grid(row=1,column=5,padx=4)
        ttk.Label(form, text="Группа (опц.)").grid(row=2,column=0,sticky="w"); self.p_group.grid(row=2,column=1,columnspan=3,sticky="we",padx=4)
        ttk.Button(form, text="Добавить", command=self._add_person).grid(row=2,column=5,sticky="e")

        btn=ttk.Frame(f); btn.pack(fill="x",padx=8,pady=4)
        ttk.Button(btn,text="Обновить",command=self._refresh_persons).pack(side="right")
        ttk.Button(btn,text="Удалить",command=self._delete_person).pack(side="right",padx=(6,0))
        ttk.Button(btn,text="Редактировать",command=self._edit_person_dialog).pack(side="right",padx=(6,0))
        ttk.Button(btn,text="Назначить в группу",command=self._assign_person_to_group).pack(side="left")

        cols=["id","ФИО","Дата рождения","Группа","Тренер","Телефон","Адрес"]
        widths=[60,200,110,170,170,120,260]
        self.tbl_persons = self.PagedSearchTable(
            self, f, cols, widths, self._person_rows,
            search_label="Поиск участника",
        )
        self.tree_persons = self.tbl_persons.tree
        self.tree_persons.bind("<Double-1>", lambda e: self._open_person_card())
        self._refresh_persons()

    def _group_options(self):
        return [self._id_label(r["group_id"], f"{r['name']} ({r['sport']}, тренер: {r['coach']})") for r in self.store.list_groups()]
    def _person_rows(self):
        return [[r["person_id"], r["fio"], r["birthdate"], r["gname"], r["coach"], r["phone"], r["address"]] for r in self.store.list_persons()]

    def _add_person(self):
        last=self.p_last.get().strip(); first=self.p_first.get().strip()
        if not last or not first: messagebox.showwarning("Данные","Фамилия и Имя обязательны"); return
        b=self.p_birth.get().strip()
        if b:
            try: dt.strptime(b,"%Y-%m-%d")
            except: messagebox.showwarning("Дата рождения","Формат: YYYY-MM-DD"); return
        gid=self._parse_id(self.p_group.get()) if self.p_group.get().strip() else None
        try: self.store.add_person(last,first,b or None,self.p_addr.get().strip() or None,self.p_phone.get().strip() or None,gid)
        except sqlite3.IntegrityError as e: messagebox.showerror("Ошибка БД", str(e)); return
        for w in (self.p_last,self.p_first,self.p_birth,self.p_addr,self.p_phone): w.delete(0,"end")
        self.p_group.set(""); self._refresh_persons(); self._refresh_groups()

    def _refresh_persons(self):
        self.tbl_persons.refresh()
    def _delete_person(self):
        sel = self.tree_persons.selection()
        if not sel:
            messagebox.showinfo("Выбор", "Выбери участника")
            return
        pid = int(self.tree_persons.item(sel[0])["values"][0])
        if not self.store.can_delete_person(pid):
            messagebox.showwarning("Нельзя удалить", "На участника есть результаты.")
            return
        if messagebox.askyesno("Подтвердите", "Удалить участника?"):
            self.store.delete_person(pid)
            self._refresh_persons()
            self._refresh_groups()

    def _edit_person_dialog(self):
        sel = self.tree_persons.selection()
        if not sel:
            messagebox.showinfo("Выбор", "Выбери участника")
            return
        pid = int(self.tree_persons.item(sel[0])["values"][0])
        p = self.store.person_raw(pid)
        if not p:
            return

        dlg = tk.Toplevel(self)
        dlg.title("Редактировать участника")
        dlg.transient(self)

        e_last  = tk.Entry(dlg, width=20); e_last.insert(0, p["last_name"])
        e_first = tk.Entry(dlg, width=20); e_first.insert(0, p["first_name"])
        e_bd    = tk.Entry(dlg, width=12); e_bd.insert(0, p["birthdate"] or "")
        e_addr  = tk.Entry(dlg, width=40); e_addr.insert(0, p["address"] or "")
        e_phone = tk.Entry(dlg, width=16); e_phone.insert(0, p["phone"] or "")
        cb_group= ttk.Combobox(dlg, width=40, values=self._group_options()); cb_group.set("")
        if p.get("group_id"):
            g = self.store._fetchone("SELECT name, sport FROM groups WHERE group_id=?", (p["group_id"],))
            if g:
                cb_group.set(self._id_label(p["group_id"], f"{g['name']} ({g['sport']})"))

        ttk.Label(dlg, text="Фамилия").grid(row=0, column=0, sticky="w", padx=6, pady=4); e_last.grid(row=0, column=1)
        ttk.Label(dlg, text="Имя").grid(row=0, column=2, sticky="w", padx=6, pady=4); e_first.grid(row=0, column=3)
        ttk.Label(dlg, text="Дата рожд. YYYY-MM-DD").grid(row=0, column=4, sticky="w", padx=6, pady=4); e_bd.grid(row=0, column=5)
        ttk.Label(dlg, text="Адрес").grid(row=1, column=0, sticky="w", padx=6, pady=4); e_addr.grid(row=1, column=1, columnspan=3, sticky="we")
        ttk.Label(dlg, text="Телефон").grid(row=1, column=4, sticky="w", padx=6, pady=4); e_phone.grid(row=1, column=5)
        ttk.Label(dlg, text="Группа (опц.)").grid(row=2, column=0, sticky="w", padx=6, pady=4); cb_group.grid(row=2, column=1, columnspan=3, sticky="we")

        def ok():
            b = e_bd.get().strip()
            if b:
                try:
                    dt.strptime(b, "%Y-%m-%d")
                except Exception:
                    messagebox.showwarning("Дата рождения", "Формат: YYYY-MM-DD")
                    return
            gid = self._parse_id(cb_group.get()) if cb_group.get().strip() else None
            self.store.edit_person(
                pid,
                e_last.get().strip(),
                e_first.get().strip(),
                b or None,
                e_addr.get().strip() or None,
                e_phone.get().strip() or None,
                gid
            )
            dlg.destroy()
            self._refresh_persons()
            self._refresh_groups()

        ttk.Button(dlg, text="Сохранить", command=ok).grid(row=3, column=5, sticky="e", padx=6, pady=8)
        dlg.grab_set()
        self.wait_window(dlg)

    def _assign_person_to_group(self):
        sel = self.tree_persons.selection()
        if not sel:
            messagebox.showinfo("Выбор", "Сначала выбери участника")
            return

        pid = int(self.tree_persons.item(sel[0])["values"][0])
        p = self.store.person_raw(pid)
        if not p:
            return

        dlg = tk.Toplevel(self)
        dlg.title("Назначить участника в группу")
        dlg.transient(self)

        ttk.Label(dlg, text=f"{p['last_name']} {p['first_name']}").grid(row=0, column=0, columnspan=2, sticky="w", padx=8, pady=(8,4))
        ttk.Label(dlg, text="Группа:").grid(row=1, column=0, sticky="w", padx=8, pady=4)

        cb = ttk.Combobox(dlg, width=50, values=self._group_options())
        cb.grid(row=1, column=1, sticky="we", padx=8, pady=4)

        # проставим текущее значение, если есть
        if p.get("group_id"):
            g = self.store._fetchone("SELECT name, sport FROM groups WHERE group_id=?", (p["group_id"],))
            if g:
                cb.set(self._id_label(p["group_id"], f"{g['name']} ({g['sport']}, тренер: —)"))

        def ok():
            gid = self._parse_id(cb.get()) if cb.get().strip() else None
            self.store.edit_person(
                pid,
                p["last_name"],
                p["first_name"],
                p["birthdate"],
                p["address"],
                p["phone"],
                gid
            )
            dlg.destroy()
            self._refresh_persons()
            self._refresh_groups()

        def clear_group():
            self.store.edit_person(
                pid,
                p["last_name"],
                p["first_name"],
                p["birthdate"],
                p["address"],
                p["phone"],
                None
            )
            dlg.destroy()
            self._refresh_persons()
            self._refresh_groups()

        btns = ttk.Frame(dlg); btns.grid(row=2, column=0, columnspan=2, sticky="e", padx=8, pady=(8,8))
        ttk.Button(btns, text="Убрать из группы", command=clear_group).pack(side="left", padx=(0,8))
        ttk.Button(btns, text="Сохранить", command=ok).pack(side="left")

        dlg.grab_set()
        self.wait_window(dlg)

    # -------- Тренеры --------
    def _tab_coaches(self):
        f=ttk.Frame(self.nb); self.nb.add(f,text="Тренеры")
        form=ttk.LabelFrame(f,text="Добавить тренера"); form.pack(fill="x",padx=8,pady=8)
        self.c_fio=tk.Entry(form,width=40); self.c_phone=tk.Entry(form,width=16)
        ttk.Label(form,text="ФИО").grid(row=0,column=0,sticky="w"); self.c_fio.grid(row=0,column=1)
        ttk.Label(form,text="Телефон").grid(row=0,column=2,sticky="w"); self.c_phone.grid(row=0,column=3)
        ttk.Button(form,text="Добавить",command=self._add_coach).grid(row=0,column=4,padx=4)

        cols=["id","ФИО","Телефон"]; widths=[60,320,160]
        self.tbl_coaches = self.PagedSearchTable(
            self, f, cols, widths, self._coach_rows,
            search_label="Поиск тренера",
        )
        self.tree_coaches = self.tbl_coaches.tree

        btn=ttk.Frame(f); btn.pack(fill="x",padx=8,pady=4)
        ttk.Button(btn,text="Обновить",command=self._refresh_coaches).pack(side="right")
        ttk.Button(btn,text="Удалить",command=self._delete_coach).pack(side="right",padx=(6,0))
        ttk.Button(btn,text="Редактировать",command=self._edit_coach_dialog).pack(side="right",padx=(6,0))

        self.tree_coaches.bind("<Double-1>", lambda e: self._open_coach_card())
        self._refresh_coaches()

    def _add_coach(self):
        fio=self.c_fio.get().strip()
        if not fio: messagebox.showwarning("Данные","Укажи ФИО"); return
        self.store.add_coach(fio, self.c_phone.get().strip() or None)
        self.c_fio.delete(0,"end"); self.c_phone.delete(0,"end"); self._refresh_coaches(); self._refresh_groups_refs()

    def _coach_rows(self): return [[r["coach_id"], r["fio"], r["phone"]] for r in self.store.list_coaches()]
    def _refresh_coaches(self): self.tbl_coaches.refresh()

    def _edit_coach_dialog(self):
        sel=self.tree_coaches.selection()
        if not sel: messagebox.showinfo("Выбор","Выбери тренера"); return
        cid=int(self.tree_coaches.item(sel[0])["values"][0])
        row=None
        for r in self.store.list_coaches():
            if r["coach_id"]==cid: row=r; break
        if not row: return
        dlg=tk.Toplevel(self); dlg.title("Редактировать тренера"); dlg.transient(self)
        e_fio=tk.Entry(dlg,width=40); e_fio.insert(0,row["fio"])
        e_phone=tk.Entry(dlg,width=16); e_phone.insert(0,row["phone"] or "")
        ttk.Label(dlg,text="ФИО").grid(row=0,column=0,sticky="w",padx=6,pady=4); e_fio.grid(row=0,column=1)
        ttk.Label(dlg,text="Телефон").grid(row=0,column=2,sticky="w",padx=6,pady=4); e_phone.grid(row=0,column=3)
        def ok():
            self.store.edit_coach(cid,e_fio.get().strip(), e_phone.get().strip() or None)
            dlg.destroy(); self._refresh_coaches(); self._refresh_groups()
        ttk.Button(dlg,text="Сохранить",command=ok).grid(row=1,column=3,sticky="e",padx=6,pady=8)
        dlg.grab_set(); self.wait_window(dlg)

    def _delete_coach(self):
        sel=self.tree_coaches.selection()
        if not sel: messagebox.showinfo("Выбор","Выбери тренера"); return
        cid=int(self.tree_coaches.item(sel[0])["values"][0])
        if not self.store.can_delete_coach(cid): messagebox.showwarning("Нельзя удалить","Тренер назначен в группе."); return
        if messagebox.askyesno("Подтвердите","Удалить тренера?"):
            self.store.delete_coach(cid); self._refresh_coaches(); self._refresh_groups_refs(); self._refresh_groups()

    # -------- Группы --------
    def _tab_groups(self):
        f=ttk.Frame(self.nb); self.nb.add(f,text="Группы")
        form=ttk.LabelFrame(f,text="Создать группу"); form.pack(fill="x",padx=8,pady=8)
        self.g_name=tk.Entry(form,width=24)
        self.g_sport=ttk.Combobox(form,values=[s for s in SPORTS if s],width=24); self.g_sport.set("Ориентирование")
        self.g_coach=ttk.Combobox(form,values=self._coach_options(),width=40); self.g_coach.set("")
        ttk.Label(form,text="Название").grid(row=0,column=0,sticky="w"); self.g_name.grid(row=0,column=1)
        ttk.Label(form,text="Вид спорта").grid(row=0,column=2,sticky="w"); self.g_sport.grid(row=0,column=3)
        ttk.Label(form,text="Тренер (опц.)").grid(row=0,column=4,sticky="w"); self.g_coach.grid(row=0,column=5)
        ttk.Button(form,text="Создать",command=self._add_group).grid(row=0,column=6,padx=4)

        cols=["id","Название","Вид спорта","Тренер","Участников"]
        self.tree_groups=ttk.Treeview(f,show="headings",columns=cols)
        for c,w in zip(cols,[60,220,160,220,100]): self.tree_groups.heading(c,text=c); self.tree_groups.column(c,width=w,anchor="w")
        # скроллы
        fr=ttk.Frame(f); fr.pack(fill="x",padx=8,pady=(4,8))
        frm=ttk.Frame(fr); frm.pack(fill="x")
        cont=ttk.Frame(f); cont.pack(fill="x",padx=8)
        tf=ttk.Frame(cont); tf.grid(row=0,column=0,sticky="nsew")
        ysb=ttk.Scrollbar(cont,orient="vertical",command=self.tree_groups.yview)
        self.tree_groups.configure(yscrollcommand=ysb.set)  # <-- фикс: правильный аргумент
        self.tree_groups.grid(in_=tf, row=0,column=0,sticky="nsew")
        ysb.grid(row=0,column=1,sticky="ns")
        tf.rowconfigure(0,weight=1); tf.columnconfigure(0,weight=1)

        frame=ttk.LabelFrame(f,text="Участники группы"); frame.pack(fill="both",expand=True,padx=8,pady=8)

        # --- ВАЖНЫЙ фикс: всё внутри gm_wrap работает на grid, и все его дети тоже на grid
        gm_wrap=ttk.Frame(frame)
        gm_wrap.pack(fill="both",expand=True,padx=4,pady=4)

        self.tree_group_members=ttk.Treeview(
            gm_wrap,  # родитель переносим в gm_wrap
            show="headings",
            columns=["id","ФИО","Дата рожд.","Телефон"]
        )
        for c,w in zip(["id","ФИО","Дата рожд.","Телефон"],[60,260,110,120]):
            self.tree_group_members.heading(c,text=c); self.tree_group_members.column(c,width=w,anchor="w")

        gm_tree=self.tree_group_members
        gm_ysb=ttk.Scrollbar(gm_wrap,orient="vertical",command=gm_tree.yview)
        gm_xsb=ttk.Scrollbar(gm_wrap,orient="horizontal",command=gm_tree.xview)
        gm_tree.configure(yscrollcommand=gm_ysb.set, xscrollcommand=gm_xsb.set)
        gm_tree.grid(row=0,column=0,sticky="nsew")
        gm_ysb.grid(row=0,column=1,sticky="ns")
        gm_xsb.grid(row=1,column=0,sticky="ew")
        gm_wrap.rowconfigure(0,weight=1); gm_wrap.columnconfigure(0,weight=1)

        btns=ttk.Frame(f); btns.pack(fill="x",padx=8,pady=4)
        ttk.Button(btns,text="Обновить",command=self._refresh_groups).pack(side="right")
        ttk.Button(btns,text="Удалить группу",command=self._delete_group).pack(side="right",padx=(6,0))
        ttk.Button(btns,text="Редактировать группу",command=self._edit_group_dialog).pack(side="right",padx=(6,0))
        ttk.Button(btns,text="Добавить участника",command=self._add_member_to_group).pack(side="left")
        ttk.Button(btns,text="Удалить участника",command=self._remove_member_from_group).pack(side="left",padx=(8,0))

        self.tree_groups.bind("<<TreeviewSelect>>", lambda e: self._refresh_group_members())
        self._refresh_groups()

    def _coach_options(self): return [self._id_label(r["coach_id"], r["fio"]) for r in self.store.list_coaches()]

    def _add_group(self):
        name=self.g_name.get().strip()
        if not name: messagebox.showwarning("Данные","Укажи название группы"); return
        coach_id=self._parse_id(self.g_coach.get()) if self.g_coach.get().strip() else None
        try: self.store.add_group(name, self.g_sport.get().strip() or "Ориентирование", coach_id)
        except sqlite3.IntegrityError as e: messagebox.showerror("Ошибка БД", str(e)); return
        self.g_name.delete(0,"end"); self.g_coach.set(""); self._refresh_groups(); self._refresh_groups_refs()

    def _group_rows(self): return [[r["group_id"],r["name"],r["sport"],r["coach"],r["members"]] for r in self.store.list_groups()]
    def _refresh_groups(self):
        self._fill_tree(self.tree_groups, self._group_rows()); self._refresh_group_members()
    def _refresh_groups_refs(self):
        self.p_group.config(values=self._group_options()); self.g_coach.config(values=self._coach_options()); self._refresh_result_refs()
    def _current_group_id(self):
        sel=self.tree_groups.selection(); return int(self.tree_groups.item(sel[0])["values"][0]) if sel else None

    def _refresh_group_members(self):
        gid=self._current_group_id(); rows=[]
        if gid:
            _,members=self.store.group_info(gid)
            for m in members: rows.append([m["person_id"],m["fio"],m["birthdate"],m["phone"]])
        self._fill_tree(self.tree_group_members, rows)

    def _choose_person_dialog(self):
        people=self.store._fetchall("SELECT person_id, last_name||' '||first_name AS fio FROM persons ORDER BY last_name, first_name")
        opts=[self._id_label(p["person_id"], p["fio"]) for p in people]
        dlg=tk.Toplevel(self); dlg.title("Выбор участника"); dlg.transient(self)
        cb=ttk.Combobox(dlg,values=opts,width=50); cb.pack(padx=8,pady=8); cb.focus()
        chosen={"id":None}
        def ok(): chosen["id"]=self._parse_id(cb.get()); dlg.destroy()
        ttk.Button(dlg,text="OK",command=ok).pack(pady=8); dlg.grab_set(); self.wait_window(dlg); return chosen["id"]

    def _add_member_to_group(self):
        gid=self._current_group_id()
        if not gid: messagebox.showinfo("Группа","Выбери группу"); return
        pid=self._choose_person_dialog()
        if not pid: return
        r=self.store.person_raw(pid)
        self.store.edit_person(pid,r["last_name"],r["first_name"],r["birthdate"],r["address"],r["phone"],gid)
        self._refresh_groups(); self._refresh_persons()

    def _remove_member_from_group(self):
        gid=self._current_group_id()
        if not gid: return
        sel=self.tree_group_members.selection()
        if not sel: messagebox.showinfo("Выбор","Выбери участника"); return
        pid=int(self.tree_group_members.item(sel[0])["values"][0])
        r=self.store.person_raw(pid)
        self.store.edit_person(pid,r["last_name"],r["first_name"],r["birthdate"],r["address"],r["phone"],None)
        self._refresh_groups(); self._refresh_persons()

    def _edit_group_dialog(self):
        gid=self._current_group_id()
        if not gid: messagebox.showinfo("Выбор","Выбери группу"); return
        g=self.store._fetchone("SELECT * FROM groups WHERE group_id=?", (gid,))
        dlg=tk.Toplevel(self); dlg.title("Редактировать группу"); dlg.transient(self)
        e_name=tk.Entry(dlg,width=24); e_name.insert(0,g["name"])
        cb_sport=ttk.Combobox(dlg,values=[s for s in SPORTS if s],width=24); cb_sport.set(g["sport"])
        cb_coach=ttk.Combobox(dlg,values=self._coach_options(),width=40)
        if g.get("coach_id"):
            c=self.store._fetchone("SELECT fio FROM coaches WHERE coach_id=?", (g["coach_id"],))
            if c: cb_coach.set(self._id_label(g["coach_id"], c["fio"]))
        ttk.Label(dlg,text="Название").grid(row=0,column=0,sticky="w",padx=6,pady=4); e_name.grid(row=0,column=1,padx=6,pady=4)
        ttk.Label(dlg,text="Вид спорта").grid(row=0,column=2,sticky="w",padx=6,pady=4); cb_sport.grid(row=0,column=3,padx=6,pady=4)
        ttk.Label(dlg,text="Тренер").grid(row=1,column=0,sticky="w",padx=6,pady=4); cb_coach.grid(row=1,column=1,columnspan=3,sticky="we",padx=6,pady=4)
        def ok():
            name=e_name.get().strip()
            if not name: messagebox.showwarning("Данные","Название обязательно"); return
            coach_id=self._parse_id(cb_coach.get()) if cb_coach.get().strip() else None
            try: self.store.edit_group(gid,name,cb_sport.get().strip() or "Ориентирование",coach_id)
            except sqlite3.IntegrityError as e: messagebox.showerror("Ошибка БД", str(e)); return
            dlg.destroy(); self._refresh_groups(); self._refresh_groups_refs(); self._refresh_persons()
        ttk.Button(dlg,text="Сохранить",command=ok).grid(row=2,column=3,sticky="e",padx=6,pady=8)
        dlg.grab_set(); self.wait_window(dlg)

    def _delete_group(self):
        gid=self._current_group_id()
        if not gid: messagebox.showinfo("Выбор","Выбери группу"); return
        if not self.store.can_delete_group(gid): messagebox.showwarning("Нельзя удалить","В группе есть участники."); return
        if messagebox.askyesno("Подтвердите","Удалить группу?"):
            self.store.delete_group(gid); self._refresh_groups(); self._refresh_groups_refs(); self._refresh_persons()

    # -------- Соревнования --------
    def _tab_events(self):
        f=ttk.Frame(self.nb); self.nb.add(f,text="Соревнования")
        form=ttk.LabelFrame(f,text="Добавить соревнование"); form.pack(fill="x",padx=8,pady=8)
        self.e_name=tk.Entry(form,width=32); self.e_date=tk.Entry(form,width=12)
        self.e_level=ttk.Combobox(form,values=[l for l in LEVELS if l],width=16); self.e_level.set("Район")
        self.e_line=ttk.Combobox(form,values=[l for l in LINES if l],width=16); self.e_line.set("Образование")
        self.e_sport=ttk.Combobox(form,values=[s for s in SPORTS if s],width=24); self.e_sport.set("Ориентирование")
        self.e_loc=tk.Entry(form,width=20); self.e_total=tk.Entry(form,width=8)
        ttk.Label(form,text="Название").grid(row=0,column=0,sticky="w"); self.e_name.grid(row=0,column=1)
        ttk.Label(form,text="Дата (YYYY-MM-DD)").grid(row=0,column=2,sticky="w"); self.e_date.grid(row=0,column=3)
        ttk.Label(form,text="Уровень").grid(row=0,column=4,sticky="w"); self.e_level.grid(row=0,column=5)
        ttk.Label(form,text="Линия").grid(row=1,column=0,sticky="w"); self.e_line.grid(row=1,column=1)
        ttk.Label(form,text="Вид спорта").grid(row=1,column=2,sticky="w"); self.e_sport.grid(row=1,column=3)
        ttk.Label(form,text="Локация").grid(row=1,column=4,sticky="w"); self.e_loc.grid(row=1,column=5)
        ttk.Label(form,text="Всего участников").grid(row=0,column=6,sticky="w"); self.e_total.grid(row=0,column=7)
        ttk.Button(form,text="Добавить",command=self._add_event).grid(row=1,column=7,sticky="e")

        cols=["id","Дата","Название","Уровень","Линия","Вид спорта","Локация","Всего","Наших"]
        widths=[60,90,240,110,120,170,150,70,70]
        def _apply_evt_tags(tree, iid, vals):  # без подсветки
            pass
        self.tbl_events = self.PagedSearchTable(
            self, f, cols, widths, self._event_rows,
            apply_tags_fn=_apply_evt_tags, search_label="Поиск соревнования"
        )
        self.tree_events = self.tbl_events.tree

        btn=ttk.Frame(f); btn.pack(fill="x",padx=8,pady=4)
        ttk.Button(btn,text="Обновить",command=self._refresh_events).pack(side="right")
        ttk.Button(btn,text="Удалить",command=self._delete_event).pack(side="right",padx=(6,0))
        ttk.Button(btn,text="Редактировать",command=self._edit_event_dialog).pack(side="right",padx=(6,0))
        self.tree_events.bind("<Double-1>", lambda e: self._open_event_card())
        self._refresh_events()

    def _add_event(self):
        name=self.e_name.get().strip(); date=self.e_date.get().strip()
        if not name or not date: messagebox.showwarning("Данные","Название и дата обязательны"); return
        try: dt.strptime(date,"%Y-%m-%d")
        except: messagebox.showwarning("Дата","Формат: YYYY-MM-DD"); return
        total=self.e_total.get().strip(); total=int(total) if total.isdigit() else None
        try: self.store.add_event(name,date,self.e_level.get().strip(),self.e_line.get().strip(),
                                  self.e_sport.get().strip(), self.e_loc.get().strip(), total)
        except sqlite3.IntegrityError as e: messagebox.showerror("Ошибка БД", str(e)); return
        for w in (self.e_name,self.e_date,self.e_loc,self.e_total): w.delete(0,"end"); self._refresh_events()

    def _event_rows(self):
        return [[r["event_id"],r["date"],r["name"],r["level"],r["line"],r["sport"],r["location"],r["total_count"],r["ours"]]
                for r in self.store.list_events()]

    def _refresh_events(self): self.tbl_events.refresh()

    def _selected_event_id(self):
        sel=self.tree_events.selection(); return int(self.tree_events.item(sel[0])["values"][0]) if sel else None

    def _edit_event_dialog(self):
        eid=self._selected_event_id()
        if not eid: messagebox.showinfo("Выбор","Выбери соревнование"); return
        e=self.store.event_raw(eid)
        dlg=tk.Toplevel(self); dlg.title("Редактировать соревнование"); dlg.transient(self)
        e_name=tk.Entry(dlg,width=32); e_name.insert(0,e["name"])
        e_date=tk.Entry(dlg,width=12); e_date.insert(0,e["date"])
        cb_lvl=ttk.Combobox(dlg,values=[l for l in LEVELS if l],width=16); cb_lvl.set(e["level"])
        cb_line=ttk.Combobox(dlg,values=[l for l in LINES if l],width=16); cb_line.set(e["line"])
        cb_spr=ttk.Combobox(dlg,values=[s for s in SPORTS if s],width=24); cb_spr.set(e["sport"])
        e_loc=tk.Entry(dlg,width=20); e_loc.insert(0,e["location"] or "")
        e_tot=tk.Entry(dlg,width=8);  e_tot.insert(0, "" if e["total_count"] is None else str(e["total_count"]))
        ttk.Label(dlg,text="Название").grid(row=0,column=0,sticky="w",padx=6,pady=4); e_name.grid(row=0,column=1)
        ttk.Label(dlg,text="Дата YYYY-MM-DD").grid(row=0,column=2,sticky="w",padx=6,pady=4); e_date.grid(row=0,column=3)
        ttk.Label(dlg,text="Уровень").grid(row=1,column=0,sticky="w",padx=6,pady=4); cb_lvl.grid(row=1,column=1)
        ttk.Label(dlg,text="Линия").grid(row=1,column=2,sticky="w",padx=6,pady=4); cb_line.grid(row=1,column=3)
        ttk.Label(dlg,text="Вид спорта").grid(row=2,column=0,sticky="w",padx=6,pady=4); cb_spr.grid(row=2,column=1)
        ttk.Label(dlg,text="Локация").grid(row=2,column=2,sticky="w",padx=6,pady=4); e_loc.grid(row=2,column=3)
        ttk.Label(dlg,text="Всего участников").grid(row=0,column=4,sticky="w",padx=6,pady=4); e_tot.grid(row=0,column=5)
        def ok():
            try: dt.strptime(e_date.get().strip(), "%Y-%m-%d")
            except: messagebox.showwarning("Дата","Формат: YYYY-MM-DD"); return
            total=e_tot.get().strip(); total=int(total) if total.isdigit() else None
            try: self.store.edit_event(eid,e_name.get().strip(),e_date.get().strip(),cb_lvl.get().strip(),cb_line.get().strip(),
                                       cb_spr.get().strip(), e_loc.get().strip() or None, total)
            except sqlite3.IntegrityError as e: messagebox.showerror("Ошибка БД", str(e)); return
            dlg.destroy(); self._refresh_events(); self._refresh_results()
        ttk.Button(dlg,text="Сохранить",command=ok).grid(row=3,column=5,sticky="e",padx=6,pady=8)
        dlg.grab_set(); self.wait_window(dlg)

    def _delete_event(self):
        eid=self._selected_event_id()
        if not eid: messagebox.showinfo("Выбор","Выбери соревнование"); return
        if not self.store.can_delete_event(eid): messagebox.showwarning("Нельзя удалить","Есть результаты."); return
        if messagebox.askyesno("Подтвердите","Удалить соревнование?"):
            self.store.delete_event(eid); self._refresh_events(); self._refresh_results()

    # -------- Результаты --------
    def _tab_results(self):
        f=ttk.Frame(self.nb); self.nb.add(f,text="Результаты")
        form=ttk.LabelFrame(f,text="Добавить результат"); form.pack(fill="x",padx=8,pady=8)
        self.r_event=ttk.Combobox(form,width=60,values=self._event_options()); self.r_event.set("")
        self.r_person=ttk.Combobox(form,width=50,values=self._person_options()); self.r_person.set("")
        self.r_cat=tk.Entry(form,width=10); self.r_place=tk.Entry(form,width=6)
        self.r_medal=ttk.Combobox(form,values=MEDALS,width=8); self.r_medal.set("")
        self.r_note=tk.Entry(form,width=20)
        ttk.Label(form,text="Соревнование").grid(row=0,column=0,sticky="w"); self.r_event.grid(row=0,column=1)
        ttk.Label(form,text="Участник").grid(row=0,column=2,sticky="w"); self.r_person.grid(row=0,column=3)
        ttk.Label(form,text="Категория").grid(row=1,column=0,sticky="w"); self.r_cat.grid(row=1,column=1)
        ttk.Label(form,text="Место").grid(row=1,column=2,sticky="w"); self.r_place.grid(row=1,column=3)
        ttk.Label(form,text="Медаль").grid(row=1,column=4,sticky="w"); self.r_medal.grid(row=1,column=5)
        ttk.Label(form,text="Прим.").grid(row=1,column=6,sticky="w"); self.r_note.grid(row=1,column=7)
        ttk.Button(form,text="Добавить",command=self._add_result).grid(row=0,column=7,sticky="e")

        cols=["id","Дата","Соревнование","Участник","Категория","Место","Медаль","Примечание"]
        widths=[60,90,240,220,90,60,80,200]
        def _apply_res_tags(tree, iid, vals):
            # vals индексы: 5 = place, 6 = medal
            self._init_all_tags(tree)
            self._apply_place_tag(tree, iid, vals[5])
            self._apply_medal_tag(tree, iid, vals[6])

        self.tbl_results = self.PagedSearchTable(
            self, f, cols, widths, self._result_rows,
            apply_tags_fn=_apply_res_tags, search_label="Поиск по результатам"
        )
        self.tree_results = self.tbl_results.tree

        btn=ttk.Frame(f); btn.pack(fill="x",padx=8,pady=4)
        ttk.Button(btn,text="Обновить списки",command=self._refresh_result_refs).pack(side="left")
        ttk.Button(btn,text="Обновить таблицу",command=self._refresh_results).pack(side="left",padx=(6,0))
        ttk.Button(btn,text="Удалить",command=self._delete_result).pack(side="right")
        ttk.Button(btn,text="Редактировать",command=self._edit_result_dialog).pack(side="right",padx=(6,0))
        self._refresh_results()

    def _event_options(self):
        return [self._id_label(r["event_id"], f"{r['date']} — {r['name']}") for r in self.store.list_events()]

    def _person_options(self):
        people=self.store._fetchall("SELECT person_id, last_name||' '||first_name AS fio FROM persons ORDER BY last_name, first_name")
        return [self._id_label(p["person_id"], p["fio"]) for p in people]

    def _refresh_result_refs(self):
        self.r_event.config(values=self._event_options()); self.r_person.config(values=self._person_options())

    def _add_result(self):
        eid=self._parse_id(self.r_event.get()); pid=self._parse_id(self.r_person.get())
        if not eid or not pid: messagebox.showwarning("Выбор","Выбери соревнование и участника"); return
        cat=(self.r_cat.get().strip() or '')
        ptxt=self.r_place.get().strip(); place=int(ptxt) if ptxt.isdigit() else None
        try: self.store.add_result(eid,pid,cat,place,self.r_medal.get().strip() or "", self.r_note.get().strip() or None)
        except sqlite3.IntegrityError as e: messagebox.showerror("Ошибка БД", str(e)); return
        for w in (self.r_cat,self.r_place,self.r_note): w.delete(0,"end"); self.r_medal.set("")
        self._refresh_results(); self._refresh_events()

    def _result_rows(self):
        return [[r["result_id"],r["date"],r["event_name"],r["fio"],r["category"],r["place"],r["medal"],r["note"]]
                for r in self.store.list_results()]

    def _refresh_results(self):
        self.tbl_results.refresh()

    def _selected_result_id(self):
        sel=self.tree_results.selection(); return int(self.tree_results.item(sel[0])["values"][0]) if sel else None

    def _delete_result(self):
        rid = self._selected_result_id()
        if not rid:
            messagebox.showinfo("Выбор", "Выбери результат")
            return
        if messagebox.askyesno("Подтвердите", "Удалить результат?"):
            self.store.delete_result(rid)
            self._refresh_results()
            self._refresh_events()

    def _edit_result_dialog(self):
        rid=self._selected_result_id()
        if not rid: messagebox.showinfo("Выбор","Выбери результат"); return
        r=self.store.result_raw(rid)
        dlg=tk.Toplevel(self); dlg.title("Редактировать результат"); dlg.transient(self)
        cb_event=ttk.Combobox(dlg,values=self._event_options(),width=60)
        cb_person=ttk.Combobox(dlg,values=self._person_options(),width=50)
        ev=self.store.event_raw(r["event_id"]); per=self.store.person_raw(r["person_id"])
        if ev:  cb_event.set(self._id_label(r["event_id"], f"{ev['date']} — {ev['name']}"))
        if per: cb_person.set(self._id_label(r["person_id"], f"{per['last_name']} {per['first_name']}"))
        e_cat=tk.Entry(dlg,width=10); e_cat.insert(0, r["category"] or "")
        e_place=tk.Entry(dlg,width=6); e_place.insert(0, "" if r["place"] is None else str(r["place"]))
        cb_medal=ttk.Combobox(dlg,values=MEDALS,width=8); cb_medal.set(r["medal"] or "")
        e_note=tk.Entry(dlg,width=20); e_note.insert(0, r["note"] or "")
        ttk.Label(dlg,text="Соревнование").grid(row=0,column=0,sticky="w",padx=6,pady=4); cb_event.grid(row=0,column=1)
        ttk.Label(dlg,text="Участник").grid(row=0,column=2,sticky="w",padx=6,pady=4); cb_person.grid(row=0,column=3)
        ttk.Label(dlg,text="Категория").grid(row=1,column=0,sticky="w",padx=6,pady=4); e_cat.grid(row=1,column=1)
        ttk.Label(dlg,text="Место").grid(row=1,column=2,sticky="w",padx=6,pady=4); e_place.grid(row=1,column=3)
        ttk.Label(dlg,text="Медаль").grid(row=1,column=4,sticky="w",padx=6,pady=4); cb_medal.grid(row=1,column=5)
        ttk.Label(dlg,text="Прим.").grid(row=1,column=6,sticky="w",padx=6,pady=4); e_note.grid(row=1,column=7)
        def ok():
            eid=self._parse_id(cb_event.get()); pid=self._parse_id(cb_person.get())
            if not eid or not pid: messagebox.showwarning("Выбор","Выберите соревнование и участника"); return
            ptxt=e_place.get().strip(); place=int(ptxt) if ptxt.isdigit() else None
            try: self.store.edit_result(rid,eid,pid,(e_cat.get().strip() or ''),place,cb_medal.get().strip() or "", e_note.get().strip() or None)
            except sqlite3.IntegrityError as e: messagebox.showerror("Ошибка БД", str(e)); return
            dlg.destroy(); self._refresh_results(); self._refresh_events()
        ttk.Button(dlg,text="Сохранить",command=ok).grid(row=2,column=7,sticky="e",padx=6,pady=8)
        dlg.grab_set(); self.wait_window(dlg)

    # -------- Отчёты + фильтры --------
    def _tab_reports(self):
        f=ttk.Frame(self.nb); self.nb.add(f,text="Отчёты")
        # ФИЛЬТРЫ
        fl=ttk.LabelFrame(f,text="Фильтры"); fl.pack(fill="x",padx=8,pady=6)
        self.f_from=tk.Entry(fl,width=12); self.f_to=tk.Entry(fl,width=12)
        self.f_sport=ttk.Combobox(fl,values=SPORTS,width=22); self.f_sport.set("")
        self.f_line=ttk.Combobox(fl,values=LINES,width=18);  self.f_line.set("")
        self.f_level=ttk.Combobox(fl,values=LEVELS,width=18); self.f_level.set("")
        ttk.Label(fl,text="С даты (YYYY-MM-DD)").pack(side="left",padx=(6,4)); self.f_from.pack(side="left")
        ttk.Label(fl,text="по").pack(side="left",padx=4); self.f_to.pack(side="left")
        ttk.Label(fl,text="Вид спорта").pack(side="left",padx=(10,4)); self.f_sport.pack(side="left")
        ttk.Label(fl,text="Линия").pack(side="left",padx=(10,4)); self.f_line.pack(side="left")
        ttk.Label(fl,text="Уровень").pack(side="left",padx=(10,4)); self.f_level.pack(side="left")
        ttk.Button(fl,text="Применить",command=self._apply_filters_and_refresh).pack(side="right",padx=6)
        ttk.Button(fl,text="Сброс",command=self._reset_filters).pack(side="right")

        # Кнопки отчётов
        top=ttk.Frame(f); top.pack(fill="x",padx=8,pady=6)
        ttk.Button(top,text="Сводка по медалям",command=self._report_medals).pack(side="left")
        ttk.Button(top,text="Соревнования по уровням/линиям",command=self._report_events_breakdown).pack(side="left",padx=8)
        ttk.Button(top,text="Итоги по тренерам",command=self._report_coaches).pack(side="left")

        ttk.Separator(f,orient="horizontal").pack(fill="x",padx=8,pady=6)

        adv=ttk.Frame(f); adv.pack(fill="x",padx=8,pady=6)
        ttk.Button(adv,text="Отчёт по участнику…",command=self._report_person_dialog).pack(side="left")
        ttk.Button(adv,text="Отчёт по группе…",command=self._report_group_dialog).pack(side="left",padx=6)
        ttk.Button(adv,text="Динамика по годам",command=self._report_yearly).pack(side="left",padx=6)

        exp=ttk.Frame(f); exp.pack(fill="x",padx=8,pady=6)
        ttk.Button(exp,text="Экспорт отчёта → TXT",command=self._export_report_txt).pack(side="left")
        ttk.Button(exp,text="Экспорт отчёта → CSV",command=self._export_report_csv).pack(side="left",padx=6)

        self.txt=tk.Text(f,wrap="word",height=24); self.txt.pack(fill="both",expand=True,padx=8,pady=8)
        self._write_report("Задайте фильтр (по желанию) и выберите отчёт.")

    def _filters(self):
        def date_or_empty(s):
            s=s.strip()
            if not s: return ""
            try: dt.strptime(s,"%Y-%m-%d"); return s
            except: return ""
        return {
            "date_from": date_or_empty(self.f_from.get()),
            "date_to":   date_or_empty(self.f_to.get()),
            "sport": self.f_sport.get().strip() or "",
            "line":  self.f_line.get().strip() or "",
            "level": self.f_level.get().strip() or "",
        }

    def _reset_filters(self):
        for w in (self.f_from,self.f_to): w.delete(0,"end")
        self.f_sport.set(""); self.f_line.set(""); self.f_level.set("")
        self._write_report("Фильтр сброшен. Выберите отчёт.")

    def _apply_filters_and_refresh(self):
        self._write_report("Фильтр применён. Теперь выберите отчёт заново.")

    def _write_report(self, text):
        self.current_report_lines = text.splitlines()
        self.txt.delete("1.0","end"); self.txt.insert("1.0", text)

    # ---- отчёты (текст в поле)
    def _report_medals(self):
        g,s,b=self.store.medals_summary(self._filters())
        self._write_report(f"Медальный зачёт (с учётом фильтра):\n  Золото: {g}\n  Серебро: {s}\n  Бронза: {b}\n  Всего: {g+s+b}\n")

    def _report_events_breakdown(self):
        by_level,by_line=self.store.events_breakdown(self._filters())
        lines=["Соревнования по уровням:"]+[f"  {k}: {by_level.get(k,0)}" for k in [x for x in LEVELS if x]]
        lines+=["","Соревнования по линиям:"]+[f"  {k}: {by_line.get(k,0)}" for k in [x for x in LINES if x]]
        self._write_report("\n".join(lines))

    def _report_coaches(self):
        rows = self.store.medals_by_coach(self._filters())
        if not rows:
            self._write_report("Нет данных по тренерам в рамках фильтра.")
            return
        rows.sort(key=lambda r: (r["g"], r["s"], r["b"], r["starts"]), reverse=True)
        lines = ["Итоги по тренерам:"]
        for r in rows:
            lines.append(
                f"  {r['fio']}: золото {r['g']}, серебро {r['s']}, бронза {r['b']}, "
                f"стартов {r['starts']}, соревнований {r['events']}, участников {r['athletes']}"
            )
        self._write_report("\n".join(lines))

    # ---- расширенные отчёты (диалоги)
    def _report_person_dialog(self):
        people=self.store._fetchall("SELECT person_id, last_name||' '||first_name AS fio FROM persons ORDER BY last_name, first_name")
        if not people: messagebox.showinfo("Нет данных","Сначала добавьте участников."); return
        opts=[self._id_label(p["person_id"], p["fio"]) for p in people]
        dlg=tk.Toplevel(self); dlg.title("Отчёт по участнику"); dlg.transient(self)
        cb=ttk.Combobox(dlg,values=opts,width=50); cb.pack(padx=8,pady=8); cb.focus()
        def ok():
            pid=self._parse_id(cb.get())
            dlg.destroy()
            if pid: self._report_person(pid)
        ttk.Button(dlg,text="OK",command=ok).pack(pady=8); dlg.grab_set(); self.wait_window(dlg)

    def _report_person(self, pid):
        rows=self.store.person_report(pid, self._filters())
        pers=self.store.person_raw(pid)
        header=f"Участник: {pers['last_name']} {pers['first_name']}\n"
        if not rows:
            self._write_report(header + "Нет стартов в рамках фильтра.")
            return
        lines=[header, "Старты:"]
        prize=0
        for r in rows:
            medal=r["medal"]; place=r["place"]; pm = f", место {place}" if place else ""
            if medal in ("gold","silver","bronze") or (place and int(place)<=3): prize+=1
            lines.append(f"  {r['date']} — {r['name']} ({r['level']}, {r['line']}, {r['sport']}) — {r['category']}{pm} {medal or ''}".rstrip())
        lines.append("")
        lines.append(f"Итого стартов: {len(rows)}; призовых: {prize}; доля призовых: {round(prize*100/len(rows),1)}%")
        self._write_report("\n".join(lines))

    def _report_group_dialog(self):
        groups=self.store.list_groups()
        if not groups: messagebox.showinfo("Нет данных","Сначала создайте группы."); return
        opts=[self._id_label(g["group_id"], f"{g['name']} ({g['sport']})") for g in groups]
        dlg=tk.Toplevel(self); dlg.title("Отчёт по группе"); dlg.transient(self)
        cb=ttk.Combobox(dlg,values=opts,width=50); cb.pack(padx=8,pady=8); cb.focus()
        def ok():
            gid=self._parse_id(cb.get())
            dlg.destroy()
            if gid: self._report_group(gid)
        ttk.Button(dlg,text="OK",command=ok).pack(pady=8); dlg.grab_set(); self.wait_window(dlg)

    def _report_group(self, gid):
        g, members = self.store.group_info(gid)
        rows = self.store.group_report(gid, self._filters())
        lines = [f"Группа: {g['name']} ({g['sport']})", ""]
        lines.append("Состав:")
        if members:
            for m in members:
                lines.append(f"  {m['fio']} ({m['birthdate'] or '—'})")
        else:
            lines.append("  —")
        lines.append("")
        lines.append("Результаты:")
        if not rows:
            lines.append("  Нет данных в рамках фильтра.")
        else:
            prize = 0
            for r in rows:
                pm = f", место {r['place']}" if r["place"] else ""
                if (r["medal"] in ("gold", "silver", "bronze")) or (r["place"] and int(r["place"]) <= 3):
                    prize += 1
                lines.append(f"  {r['date']} — {r['name']} — {r['fio']} — {r['category']}{pm} {r['medал'] if 'медал' in r else (r['medal'] or '')}".rstrip())
            lines.append("")
            lines.append(f"Итого результатов: {len(rows)}; призовых: {prize}")
        self._write_report("\n".join(lines))

    def _report_yearly(self):
        rows=self.store.yearly_dynamics(self._filters())
        if not rows:
            self._write_report("Нет данных для построения динамики.")
            return
        lines=["Динамика по годам:"]
        for r in rows:
            lines.append(f"  {r['year']}: стартов {r['events']}, медалей {r['total_medals']} (зол {r['gold']}, сер {r['silver']}, бронз {r['bronze']})")
        self._write_report("\n".join(lines))

    # --- Экспорт текущего отчёта
    def _export_report_txt(self):
        path = filedialog.asksaveasfilename(
            title="Сохранить отчёт (TXT)", initialfile=f"report_{self._timestamp()}.txt",
            defaultextension=".txt", filetypes=[("TXT","*.txt")]
        )
        if not path: return
        with io.open(path,"w",encoding="utf-8") as f:
            f.write(self.txt.get("1.0","end").rstrip())
        messagebox.showinfo("Экспорт", f"Сохранено: {path}")

    def _export_report_csv(self):
        path = filedialog.asksaveasfilename(
            title="Сохранить отчёт (CSV)", initialfile=f"report_{self._timestamp()}.csv",
            defaultextension=".csv", filetypes=[("CSV","*.csv")]
        )
        if not path: return
        text = self.txt.get("1.0","end").rstrip().splitlines()
        with io.open(path,"w",encoding="utf-8-sig",newline="") as f:
            w = csv.writer(f, delimiter=';')
            for line in text:
                if ": " in line:
                    k,v = line.split(": ",1)
                    w.writerow([k.strip(), v.strip()])
                else:
                    w.writerow([line])
        messagebox.showinfo("Экспорт", f"Сохранено: {path}")

    # ---------- Карточки (участник / тренер / соревнование) ----------
    def _mini_filter_frame(self, parent):
        fr = ttk.Frame(parent)
        e_from = tk.Entry(fr, width=12)
        e_to   = tk.Entry(fr, width=12)
        cb_sport = ttk.Combobox(fr, values=SPORTS, width=18); cb_sport.set("")
        cb_line  = ttk.Combobox(fr, values=LINES,  width=14); cb_line.set("")
        cb_level = ttk.Combobox(fr, values=LEVELS, width=16); cb_level.set("")
        ttk.Label(fr,text="с").pack(side="left"); e_from.pack(side="left",padx=3)
        ttk.Label(fr,text="по").pack(side="left",padx=3); e_to.pack(side="left",padx=3)
        ttk.Label(fr,text="спорт").pack(side="left",padx=(6,3)); cb_sport.pack(side="left")
        ttk.Label(fr,text="линия").pack(side="left",padx=(6,3)); cb_line.pack(side="left")
        ttk.Label(fr,text="уровень").pack(side="left",padx=(6,3)); cb_level.pack(side="left")
        return fr, e_from, e_to, cb_sport, cb_line, cb_level

    def _collect_mini_filter(self, e_from, e_to, cb_sport, cb_line, cb_level):
        def okdate(s):
            s = s.strip()
            if not s: return ""
            try: dt.strptime(s, "%Y-%m-%d"); return s
            except: return ""
        return {
            "date_from": okdate(e_from.get()),
            "date_to":   okdate(e_to.get()),
            "sport": cb_sport.get().strip() or "",
            "line":  cb_line.get().strip() or "",
            "level": cb_level.get().strip() or "",
        }

    def _open_person_card(self):
        sel = self.tree_persons.selection()
        if not sel: return
        pid = int(self.tree_persons.item(sel[0])["values"][0])
        p = self.store.person_raw(pid)

        win = tk.Toplevel(self); win.title(f"Участник — {p['last_name']} {p['first_name']}"); win.geometry("820x520"); win.transient(self)

        top = ttk.Frame(win); top.pack(fill="x", padx=8, pady=8)
        ttk.Label(top, text=f"{p['last_name']} {p['first_name']}  •  рожд.: {p.get('birthdate') or '—'}  •  тел.: {p.get('phone') or '—'}").pack(side="left")

        fl, e_from, e_to, cb_sport, cb_line, cb_level = self._mini_filter_frame(win)
        fl.pack(fill="x", padx=8)
        summary = tk.StringVar(value="—")
        ttk.Label(win, textvariable=summary).pack(anchor="w", padx=10, pady=(6,0))

        cols = ["Дата","Соревнование","Линия","Уровень","Вид спорта","Категория","Место","Медаль","Прим."]
        tree = ttk.Treeview(win, show="headings", columns=cols)
        for c,w in zip(cols,[90,240,110,100,120,110,60,70,200]):
            tree.heading(c,text=c); tree.column(c,width=w,anchor="w")
        # скроллы
        wrap = ttk.Frame(win); wrap.pack(fill="both", expand=True, padx=8, pady=8)
        ysb = ttk.Scrollbar(wrap, orient="vertical", command=tree.yview)
        xsb = ttk.Scrollbar(wrap, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        wrap.rowconfigure(0, weight=1); wrap.columnconfigure(0, weight=1)

        self._init_all_tags(tree)

        def refresh():
            flt = self._collect_mini_filter(e_from,e_to,cb_sport,cb_line,cb_level)
            rows = self.store.person_report(pid, flt)
            for i in tree.get_children(): tree.delete(i)
            for r in rows:
                vals = [r["date"], r["name"], r["line"], r["level"], r["sport"], r["category"], r["place"], r["medal"], r["note"]]
                iid = tree.insert("", "end", values=vals)
                self._apply_place_tag(tree, iid, r["place"])
                self._apply_medal_tag(tree, iid, r["medal"])
            s = self.store.person_summary(pid, flt)
            summary.set(f"Итого стартов: {s['starts']}  •  призовых: {s['prize']}  •  медали — зол: {s['gold']}, сер: {s['silver']}, бронз: {s['bronze']}")
        ttk.Button(fl, text="Применить", command=refresh).pack(side="right", padx=6)
        refresh()

    def _open_coach_card(self):
        sel = self.tree_coaches.selection()
        if not sel: return
        cid = int(self.tree_coaches.item(sel[0])["values"][0])
        c = None
        for r in self.store.list_coaches():
            if r["coach_id"]==cid: c=r; break
        if not c: return

        win = tk.Toplevel(self); win.title(f"Тренер — {c['fio']}"); win.geometry("900x560"); win.transient(self)

        top = ttk.Frame(win); top.pack(fill="x", padx=8, pady=8)
        ttk.Label(top, text=f"{c['fio']}  •  тел.: {c.get('phone') or '—'}").pack(side="left")

        fl, e_from, e_to, cb_sport, cb_line, cb_level = self._mini_filter_frame(win)
        fl.pack(fill="x", padx=8)
        summary = tk.StringVar(value="—")
        ttk.Label(win, textvariable=summary).pack(anchor="w", padx=10, pady=(6,0))

        cols = ["Дата","Соревнование","Участник","Категория","Место","Медаль","Прим."]
        tree = ttk.Treeview(win, show="headings", columns=cols)
        for c,w in zip(cols,[90,280,220,110,60,70,200]):
            tree.heading(c,text=c); tree.column(c,width=w,anchor="w")
        # скроллы
        wrap = ttk.Frame(win); wrap.pack(fill="both", expand=True, padx=8, pady=8)
        ysb = ttk.Scrollbar(wrap, orient="vertical", command=tree.yview)
        xsb = ttk.Scrollbar(wrap, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        wrap.rowconfigure(0, weight=1); wrap.columnconfigure(0, weight=1)

        self._init_all_tags(tree)

        def refresh():
            flt = self._collect_mini_filter(e_from,e_to,cb_sport,cb_line,cb_level)
            rows = self.store.coach_results(cid, flt)
            for i in tree.get_children(): tree.delete(i)
            for r in rows:
                vals = [r["date"], r["event_name"], r["fio"], r["category"], r["place"], r["medal"], r["note"]]
                iid = tree.insert("", "end", values=vals)
                self._apply_place_tag(tree, iid, r["place"])
                self._apply_medal_tag(tree, iid, r["medal"])
            s = self.store.coach_summary(cid, flt)
            summary.set(f"Стартов: {s['starts']} • Соревнований: {s['events']} • Спортсменов: {s['athletes']} • Медали — зол:{s['g']} сер:{s['s']} бронз:{s['b']}")
        ttk.Button(fl, text="Применить", command=refresh).pack(side="right", padx=6)
        refresh()

    def _open_event_card(self):
        sel = self.tree_events.selection()
        if not sel: return
        eid = int(self.tree_events.item(sel[0])["values"][0])
        e = self.store.event_raw(eid)
        win = tk.Toplevel(self); win.title(f"Соревнование — {e['date']} • {e['name']}"); win.geometry("940x560"); win.transient(self)

        ttk.Label(win, text=f"{e['date']}  —  {e['name']}  •  {e['level']} / {e['line']} • {e['sport']} • {e.get('location') or '—'}").pack(
            anchor="w", padx=10, pady=10)

        cols = ["Место","Медаль","Участник","Группа","Тренер","Категория","Прим."]
        tree = ttk.Treeview(win, show="headings", columns=cols)
        for c,w in zip(cols,[60,70,220,160,180,120,200]):
            tree.heading(c,text=c); tree.column(c,width=w,anchor="w")

        wrap = ttk.Frame(win); wrap.pack(fill="both", expand=True, padx=8, pady=8)
        ysb = ttk.Scrollbar(wrap, orient="vertical", command=tree.yview)
        xsb = ttk.Scrollbar(wrap, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        wrap.rowconfigure(0, weight=1); wrap.columnconfigure(0, weight=1)

        self._init_all_tags(tree)

        rows = self.store.event_results(eid)
        for r in rows:
            vals = [r["place"], r["medal"], r["fio"], r["gname"], r["coach"], r["category"], r["note"]]
            iid = tree.insert("", "end", values=vals)
            self._apply_place_tag(tree, iid, r["place"])
            self._apply_medal_tag(tree, iid, r["medal"])

    # -------- Импорт/Экспорт --------
    def _tab_io(self):
        f=ttk.Frame(self.nb); self.nb.add(f,text="Импорт/Экспорт")

        # Экспорт
        box=ttk.LabelFrame(f,text="Экспорт"); box.pack(fill="x",padx=8,pady=8)
        ttk.Button(box,text="Экспорт всех таблиц → CSV",command=self._export_all_csv).pack(side="left",padx=6,pady=6)
        if HAS_XLSX:
            ttk.Button(box,text="Экспорт всех таблиц → XLSX",command=self._export_xlsx).pack(side="left",padx=6,pady=6)
        else:
            ttk.Label(box,text="(для XLSX установи пакет openpyxl)").pack(side="left",padx=6)

        # Импорт
        box2=ttk.LabelFrame(f,text="Импорт"); box2.pack(fill="x",padx=8,pady=8)
        self.var_clear= tk.BooleanVar(value=False)
        ttk.Checkbutton(box2,text="Очистить таблицу перед импортом",variable=self.var_clear).pack(side="right",padx=6)
        ttk.Button(box2,text="Импорт coaches.csv",command=lambda:self._import_csv("coaches")).pack(side="left",padx=6,pady=6)
        ttk.Button(box2,text="Импорт groups.csv",command=lambda:self._import_csv("groups")).pack(side="left",padx=6)
        ttk.Button(box2,text="Импорт persons.csv",command=lambda:self._import_csv("persons")).pack(side="left",padx=6)
        ttk.Button(box2,text="Импорт events.csv",command=lambda:self._import_csv("events")).pack(side="left",padx=6)
        ttk.Button(box2,text="Импорт results.csv",command=lambda:self._import_csv("results")).pack(side="left",padx=6)
        if HAS_XLSX:
            ttk.Button(box2,text="Импорт из XLSX (все листы)",command=self._import_xlsx).pack(side="left",padx=6)

        # Шаблоны
        box3=ttk.LabelFrame(f,text="Шаблоны файлов"); box3.pack(fill="x",padx=8,pady=8)
        ttk.Button(box3,text="Создать шаблоны CSV",command=self._make_csv_templates).pack(side="left",padx=6,pady=6)
        if HAS_XLSX:
            ttk.Button(box3,text="Создать шаблон XLSX",command=self._make_xlsx_template).pack(side="left",padx=6)

        # Памятка форматов
        tips=ttk.LabelFrame(f,text="Формат колонок"); tips.pack(fill="x",padx=8,pady=8)
        txt=tk.Text(tips,height=10,wrap="word"); txt.pack(fill="x",padx=6,pady=6)
        txt.insert("1.0",
            "CSV: кодировка UTF-8 с BOM, разделитель ;\n"
            "coaches: coach_id(optional), fio, phone\n"
            "groups: group_id(optional), name, sport, coach_id(optional)\n"
            "persons: person_id(optional), last_name, first_name, birthdate(YYYY-MM-DD), address, phone, group_id(optional)\n"
            "events: event_id(optional), name, date(YYYY-MM-DD), level, line, sport, location, total_count\n"
            "results: result_id(optional), event_id, person_id, category, place, medal, note\n"
            "Важно: импорт results делай после загрузки persons и events.\n"
        )
        txt.config(state="disabled")

    # ---- экспорт всех таблиц в CSV/XLSX и импорт
    def _save_dialog(self, title, defname):
        return filedialog.asksaveasfilename(title=title, initialfile=defname, defaultextension=".csv",
                                            filetypes=[("CSV","*.csv"),("All","*.*")])
    def _open_dialog(self, title):
        return filedialog.askopenfilename(title=title, filetypes=[("CSV","*.csv"),("All","*.*")])

    def _export_all_csv(self):
        base = f"export_{self._timestamp()}"
        folder = filedialog.askdirectory(title="Куда сохранить CSV?")
        if not folder: return
        def write_csv(path, header, rows):
            with io.open(path,"w",encoding="utf-8-sig",newline="") as f:
                w = csv.writer(f, delimiter=';')
                w.writerow(header); w.writerows(rows)
        # coaches
        write_csv(os.path.join(folder,f"{base}_coaches.csv"),
                  ["coach_id","fio","phone"],
                  [[r["coach_id"],r["fio"],r["phone"]] for r in self.store.list_coaches()])
        # groups
        glist = self.store._fetchall("SELECT group_id,name,sport,coach_id FROM groups ORDER BY group_id")
        write_csv(os.path.join(folder,f"{base}_groups.csv"),
                  ["group_id","name","sport","coach_id"],
                  [[r["group_id"],r["name"],r["sport"],r["coach_id"] or ""] for r in glist])
        # persons
        plist = self.store._fetchall("SELECT person_id,last_name,first_name,birthdate,address,phone,group_id FROM persons ORDER BY person_id")
        write_csv(os.path.join(folder,f"{base}_persons.csv"),
                  ["person_id","last_name","first_name","birthdate","address","phone","group_id"],
                  [[r["person_id"],r["last_name"],r["first_name"],r["birthdate"] or "",r["address"] or "",r["phone"] or "",r["group_id"] or ""] for r in plist])
        # events
        elist = self.store._fetchall("SELECT event_id,name,date,level,line,sport,location,total_count FROM events ORDER BY date DESC, event_id DESC")
        write_csv(os.path.join(folder,f"{base}_events.csv"),
                  ["event_id","name","date","level","line","sport","location","total_count"],
                  [[r["event_id"],r["name"],r["date"],r["level"],r["line"],r["sport"],r["location"] or "",r["total_count"] or ""] for r in elist])
        # results
        rlist = self.store._fetchall("SELECT result_id,event_id,person_id,category,place,medal,note FROM results ORDER BY result_id")
        write_csv(os.path.join(folder,f"{base}_results.csv"),
                  ["result_id","event_id","person_id","category","place","medal","note"],
                  [[r["result_id"],r["event_id"],r["person_id"],r["category"] or "",r["place"] or "",r["medal"] or "",r["note"] or ""] for r in rlist])
        messagebox.showinfo("Экспорт CSV", f"Готово. Файлы сохранены в:\n{folder}")

    def _import_csv(self, table):
        path = self._open_dialog(f"Выбери {table}.csv")
        if not path: return
        clear = self.var_clear.get()
        try:
            with io.open(path,"r",encoding="utf-8-sig") as f:
                r = csv.reader(f, delimiter=';')
                header = next(r, [])
                rows = list(r)
            cur = self.store.conn.cursor()
            cur.execute("BEGIN")
            if clear:
                cur.execute(f"DELETE FROM {table}")
            if table == "coaches":
                idx = {h:i for i,h in enumerate(header)}
                for row in rows:
                    fio = row[idx.get("fio",1)].strip()
                    phone = row[idx.get("phone",2)].strip() or None
                    cidtxt = row[idx.get("coach_id",0)].strip() if "coach_id" in idx else ""
                    if cidtxt:
                        cur.execute("INSERT OR REPLACE INTO coaches(coach_id,fio,phone) VALUES(?,?,?)", (int(cidtxt), fio, phone))
                    else:
                        cur.execute("INSERT INTO coaches(fio,phone) VALUES(?,?)", (fio, phone))
            elif table == "groups":
                idx = {h:i for i,h in enumerate(header)}
                for row in rows:
                    name = row[idx.get("name",1)].strip()
                    sport= row[idx.get("sport",2)].strip() or "Ориентирование"
                    coach = row[idx.get("coach_id",3)].strip()
                    gidtxt = row[idx.get("group_id",0)].strip() if "group_id" in idx else ""
                    coach_id = int(coach) if coach else None
                    if gidtxt:
                        cur.execute("INSERT OR REPLACE INTO groups(group_id,name,sport,coach_id) VALUES(?,?,?,?)",
                                    (int(gidtxt), name, sport, coach_id))
                    else:
                        cur.execute("INSERT INTO groups(name,sport,coach_id) VALUES(?,?,?)",
                                    (name, sport, coach_id))
            elif table == "persons":
                idx = {h:i for i,h in enumerate(header)}
                for row in rows:
                    last=row[idx.get("last_name",1)].strip(); first=row[idx.get("first_name",2)].strip()
                    birth=row[idx.get("birthdate",3)].strip() or None
                    addr=row[idx.get("address",4)].strip() or None
                    phone=row[idx.get("phone",5)].strip() or None
                    gid=row[idx.get("group_id",6)].strip(); group_id=int(gid) if gid else None
                    pidtxt=row[idx.get("person_id",0)].strip() if "person_id" in idx else ""
                    if pidtxt:
                        cur.execute("""INSERT OR REPLACE INTO persons(person_id,last_name,first_name,birthdate,address,phone,group_id)
                                       VALUES(?,?,?,?,?,?,?)""",(int(pidtxt),last,first,birth,addr,phone,group_id))
                    else:
                        cur.execute("""INSERT INTO persons(last_name,first_name,birthdate,address,phone,group_id)
                                       VALUES(?,?,?,?,?,?)""",(last,first,birth,addr,phone,group_id))
            elif table == "events":
                idx = {h:i for i,h in enumerate(header)}
                for row in rows:
                    name=row[idx.get("name",1)].strip()
                    date=row[idx.get("date",2)].strip()
                    level=row[idx.get("level",3)].strip() or "Район"
                    line=row[idx.get("line",4)].strip() or "Образование"
                    sport=row[idx.get("sport",5)].strip() or "Ориентирование"
                    loc=row[idx.get("location",6)].strip() or None
                    tot=row[idx.get("total_count",7)].strip(); total=int(tot) if tot.isdigit() else None
                    eidtxt=row[idx.get("event_id",0)].strip() if "event_id" in idx else ""
                    if eidtxt:
                        cur.execute("""INSERT OR REPLACE INTO events(event_id,name,date,level,line,sport,location,total_count)
                                       VALUES(?,?,?,?,?,?,?,?)""",(int(eidtxt),name,date,level,line,sport,loc,total))
                    else:
                        cur.execute("""INSERT INTO events(name,date,level,line,sport,location,total_count)
                                       VALUES(?,?,?,?,?,?,?)""",(name,date,level,line,sport,loc,total))
            elif table == "results":
                idx = {h:i for i,h in enumerate(header)}
                for row in rows:
                    eid=int(row[idx.get("event_id",1)].strip())
                    pid=int(row[idx.get("person_id",2)].strip())
                    cat=(row[idx.get("category",3)].strip() or '')
                    plc=row[idx.get("place",4)].strip(); place=int(plc) if plc.isdigit() else None
                    medal=row[idx.get("medal",5)].strip() or ""
                    note=row[idx.get("note",6)].strip() or None
                    ridtxt=row[idx.get("result_id",0)].strip() if "result_id" in idx else ""
                    if ridtxt:
                        cur.execute("""INSERT OR REPLACE INTO results(result_id,event_id,person_id,category,place,medal,note)
                                       VALUES(?,?,?,?,?,?,?)""",(int(ridtxt),eid,pid,cat,place,medal,note))
                    else:
                        cur.execute("""INSERT OR REPLACE INTO results(event_id,person_id,category,place,medal,note)
                                       VALUES(?,?,?,?,?,?)""",(eid,pid,cat,place,medal,note))
            cur.execute("COMMIT")
        except Exception as e:
            try: cur.execute("ROLLBACK")
            except: pass
            messagebox.showerror("Импорт CSV — ошибка", str(e))
            return
        messagebox.showinfo("Импорт CSV", f"Импорт завершён: {table}")
        self._refresh_coaches(); self._refresh_groups(); self._refresh_persons(); self._refresh_events(); self._refresh_results(); self._refresh_result_refs()

    # --- XLSX (если есть openpyxl)
    def _export_xlsx(self):
        if not HAS_XLSX: return
        path = filedialog.asksaveasfilename(title="Сохранить XLSX", initialfile=f"export_{self._timestamp()}.xlsx",
                                            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not path: return
        wb = openpyxl.Workbook()
        def ws(name, header, rows):
            ws = wb.create_sheet(title=name)
            ws.append(header)
            for row in rows: ws.append(row)
        wb.remove(wb.active)
        ws("coaches", ["coach_id","fio","phone"],
           [[r["coach_id"],r["fio"],r["phone"]] for r in self.store.list_coaches()])
        glist=self.store._fetchall("SELECT group_id,name,sport,coach_id FROM groups ORDER BY group_id")
        ws("groups", ["group_id","name","sport","coach_id"],
           [[r["group_id"],r["name"],r["sport"],r["coach_id"]] for r in glist])
        plist=self.store._fetchall("SELECT person_id,last_name,first_name,birthdate,address,phone,group_id FROM persons ORDER BY person_id")
        ws("persons", ["person_id","last_name","first_name","birthdate","address","phone","group_id"],
           [[r["person_id"],r["last_name"],r["first_name"],r["birthdate"],r["address"],r["phone"],r["group_id"]] for r in plist])
        elist=self.store._fetchall("SELECT event_id,name,date,level,line,sport,location,total_count FROM events ORDER BY date DESC, event_id DESC")
        ws("events", ["event_id","name","date","level","line","sport","location","total_count"],
           [[r["event_id"],r["name"],r["date"],r["level"],r["line"],r["sport"],r["location"],r["total_count"]] for r in elist])
        rlist=self.store._fetchall("SELECT result_id,event_id,person_id,category,place,medal,note FROM results ORDER BY result_id")
        ws("results", ["result_id","event_id","person_id","category","place","medal","note"],
           [[r["result_id"],r["event_id"],r["person_id"],r["category"],r["place"],r["medal"],r["note"]] for r in rlist])
        wb.save(path)
        messagebox.showinfo("Экспорт XLSX", f"Сохранено: {path}")

    def _import_xlsx(self):
        if not HAS_XLSX: return
        path = filedialog.askopenfilename(title="Выбери XLSX", filetypes=[("Excel","*.xlsx")])
        if not path: return
        clear = self.var_clear.get()
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            cur = self.store.conn.cursor(); cur.execute("BEGIN")
            order = ["coaches","groups","persons","events","results"]
            for name in order:
                if name not in wb.sheetnames: continue
                ws = wb[name]
                rows = list(ws.values)
                if not rows: continue
                header = [str(h) if h is not None else "" for h in rows[0]]
                data = rows[1:]
                if clear:
                    cur.execute(f"DELETE FROM {name}")
                idx = {h:i for i,h in enumerate(header)}
                if name == "coaches":
                    for row in data:
                        fio = str(row[idx.get("fio",1)] or "").strip()
                        phone = str(row[idx.get("phone",2)] or "").strip() or None
                        cidtxt = str(row[idx.get("coach_id",0)] or "").strip() if "coach_id" in idx else ""
                        if cidtxt:
                            cur.execute("INSERT OR REPLACE INTO coaches(coach_id,fio,phone) VALUES(?,?,?)", (int(cidtxt), fio, phone))
                        else:
                            cur.execute("INSERT INTO coaches(fio,phone) VALUES(?,?)", (fio, phone))
                elif name == "groups":
                    for row in data:
                        namev = str(row[idx.get("name",1)] or "").strip()
                        sport = str(row[idx.get("sport",2)] or "").strip() or "Ориентирование"
                        coach = str(row[idx.get("coach_id",3)] or "").strip()
                        gidtxt = str(row[idx.get("group_id",0)] or "").strip() if "group_id" in idx else ""
                        coach_id = int(coach) if coach else None
                        if gidtxt:
                            cur.execute("INSERT OR REPLACE INTO groups(group_id,name,sport,coach_id) VALUES(?,?,?,?)",
                                        (int(gidtxt), namev, sport, coach_id))
                        else:
                            cur.execute("INSERT INTO groups(name,sport,coach_id) VALUES(?,?,?)",
                                        (namev, sport, coach_id))
                elif name == "persons":
                    for row in data:
                        last = str(row[idx.get("last_name",1)] or "").strip()
                        first= str(row[idx.get("first_name",2)] or "").strip()
                        birth= str(row[idx.get("birthdate",3)] or "").strip() or None
                        addr = str(row[idx.get("address",4)] or "").strip() or None
                        phone= str(row[idx.get("phone",5)] or "").strip() or None
                        gid  = str(row[idx.get("group_id",6)] or "").strip()
                        group_id = int(gid) if gid else None
                        pidtxt = str(row[idx.get("person_id",0)] or "").strip() if "person_id" in idx else ""
                        if pidtxt:
                            cur.execute("""INSERT OR REPLACE INTO persons(person_id,last_name,first_name,birthdate,address,phone,group_id)
                                           VALUES(?,?,?,?,?,?,?)""",(int(pidtxt),last,first,birth,addr,phone,group_id))
                        else:
                            cur.execute("""INSERT INTO persons(last_name,first_name,birthdate,address,phone,group_id)
                                           VALUES(?,?,?,?,?,?)""",(last,first,birth,addr,phone,group_id))
                elif name == "events":
                    for row in data:
                        namev = str(row[idx.get("name",1)] or "").strip()
                        date  = str(row[idx.get("date",2)] or "").strip()
                        level = str(row[idx.get("level",3)] or "").strip() or "Район"
                        line  = str(row[idx.get("line",4)] or "").strip() or "Образование"
                        sport = str(row[idx.get("sport",5)] or "").strip() or "Ориентирование"
                        loc   = str(row[idx.get("location",6)] or "").strip() or None
                        tot   = str(row[idx.get("total_count",7)] or "").strip()
                        total = int(tot) if tot.isdigit() else None
                        eidtxt= str(row[idx.get("event_id",0)] or "").strip() if "event_id" in idx else ""
                        if eidtxt:
                            cur.execute("""INSERT OR REPLACE INTO events(event_id,name,date,level,line,sport,location,total_count)
                                           VALUES(?,?,?,?,?,?,?,?)""",(int(eidtxt),namev,date,level,line,sport,loc,total))
                        else:
                            cur.execute("""INSERT INTO events(name,date,level,line,sport,location,total_count)
                                           VALUES(?,?,?,?,?,?,?)""",(namev,date,level,line,sport,loc,total))
                elif name == "results":
                    for row in data:
                        eid = int(str(row[idx.get("event_id",1)] or "0").strip())
                        pid = int(str(row[idx.get("person_id",2)] or "0").strip())
                        cat = str(row[idx.get("category",3)] or "").strip()
                        plc = str(row[idx.get("place",4)] or "").strip()
                        place = int(plc) if plc.isdigit() else None
                        medal = str(row[idx.get("medal",5)] or "").strip()
                        note  = str(row[idx.get("note",6)] or "").strip() or None
                        ridtxt= str(row[idx.get("result_id",0)] or "").strip() if "result_id" in idx else ""
                        if ridtxt:
                            cur.execute("""INSERT OR REPLACE INTO results(result_id,event_id,person_id,category,place,medal,note)
                                           VALUES(?,?,?,?,?,?,?)""",(int(ridtxt),eid,pid,cat,place,medal,note))
                        else:
                            cur.execute("""INSERT OR REPLACE INTO results(event_id,person_id,category,place,medal,note)
                                           VALUES(?,?,?,?,?,?)""",(eid,pid,cat,place,medal,note))
            cur.execute("COMMIT")
        except Exception as e:
            try: cur.execute("ROLLBACK")
            except: pass
            messagebox.showerror("Импорт XLSX — ошибка", str(e)); return
        messagebox.showinfo("Импорт XLSX", f"Импорт завершён: {os.path.basename(path)}")
        self._refresh_coaches(); self._refresh_groups(); self._refresh_persons(); self._refresh_events(); self._refresh_results(); self._refresh_result_refs()

    # ------------- шаблоны -------------
    def _make_csv_templates(self):
        folder = filedialog.askdirectory(title="Где создать шаблоны CSV?")
        if not folder: return
        def make(name, header):
            path=os.path.join(folder, f"{name}.csv")
            with io.open(path,"w",encoding="utf-8-sig",newline="") as f:
                csv.writer(f, delimiter=';').writerow(header)
        make("coaches", ["coach_id","fio","phone"])
        make("groups",  ["group_id","name","sport","coach_id"])
        make("persons", ["person_id","last_name","first_name","birthdate","address","phone","group_id"])
        make("events",  ["event_id","name","date","level","line","sport","location","total_count"])
        make("results", ["result_id","event_id","person_id","category","place","medal","note"])
        messagebox.showinfo("Шаблоны CSV", f"Созданы в папке:\n{folder}")

    def _make_xlsx_template(self):
        if not HAS_XLSX:
            messagebox.showinfo("XLSX","Установи пакет openpyxl, чтобы создавать XLSX.")
            return
        path = filedialog.asksaveasfilename(title="Сохранить шаблон XLSX", initialfile="template.xlsx",
                                            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not path: return
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        def ws(name, header):
            s=wb.create_sheet(title=name); s.append(header)
        ws("coaches", ["coach_id","fio","phone"])
        ws("groups",  ["group_id","name","sport","coach_id"])
        ws("persons", ["person_id","last_name","first_name","birthdate","address","phone","group_id"])
        ws("events",  ["event_id","name","date","level","line","sport","location","total_count"])
        ws("results", ["result_id","event_id","person_id","category","place","medal","note"])
        wb.save(path); messagebox.showinfo("Шаблон XLSX", f"Сохранено: {path}")

# ------------- run -------------
if __name__ == "__main__":
    App().mainloop()

